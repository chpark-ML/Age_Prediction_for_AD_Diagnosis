import nibabel as nib
from matplotlib import gridspec
import utils as ut
import numbers
from torch.nn import functional as F
from torch.autograd import Variable
import math
import torch
import torch.nn as nn
import shutil
import numpy as np
import os
import setting as st
import setting_2 as fst
import nibabel as nib
import matplotlib
matplotlib.use('Agg')
import torch
from sklearn import metrics
from sklearn.metrics import r2_score, mean_squared_error
from sklearn.metrics import confusion_matrix
from sklearn.utils.multiclass import unique_labels
from scipy import stats
import matplotlib.pyplot as plt
import utils
import pickle
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from collections import deque
from collections import OrderedDict
import seaborn as sns
import matplotlib.colors as mcolors
import pandas as pd
from pandas import Series, DataFrame

def ROI_Gaussian_normalization(RoI_template, data, list_RoI_mean=0, list_RoI_std=0, train = True):
    # data: [sample, channel, freq, 1]
    # (sample, channel, time)
    if len(data.shape) != 5:
        data = data[:, :, 0, :, :, :]
    if data.shape[1] != 1:  # GM only
        data = data[:, 0:1]

    n_ROI = np.unique(RoI_template).shape[0] - 1
    ##TODO : calculate ROI-wise mean, std
    if train == True:
        list_RoI_mean = []
        list_RoI_std = []
        for i in range(1, n_ROI+1):
            list_RoI_mean.append(data[:, :, (RoI_template == i)].mean())
            list_RoI_std.append(data[:, :, (RoI_template == i)].std())

    ##TODO : normalize
    for i in range(1, n_ROI + 1):
        data[:, :, (RoI_template == i)] = ((data[:, :, (RoI_template == i)] - list_RoI_mean[i - 1]) / (list_RoI_std[i - 1]))

    assert np.sum(np.isinf(data)) == 0

    return data, list_RoI_mean, list_RoI_std


class elbo_loss(nn.Module):
    def __init__(self):
        super(elbo_loss, self).__init__()

    def foward(self, input, target, t, num_batches, sample, debug=False):
        if sample:
            lps, lvps, predictions = [], [], []
            for i in range(self.samples):
                predictions.append(self.model(input, sample=sample)[t])
                lp, lv = self.logs(t)
                lps.append(lp)
                lvps.append(lv)

            # hack
            w1 = 1.e-3
            w2 = 1.e-3
            w3 = 5.e-2

            outputs = torch.stack(predictions, dim=0).to(self.device)
            log_var = w1 * torch.as_tensor(lvps, device=self.device).mean()
            log_p = w2 * torch.as_tensor(lps, device=self.device).mean()
            nll = w3 * torch.nn.functional.nll_loss(outputs.mean(0), target, reduction='sum').to(device=self.device)

            return (log_var - log_p) / num_batches + nll

        else:
            predictions = []
            for i in range(self.samples):
                pred = self.model(input, sample=False)[t]
                predictions.append(pred)

            # hack
            # w1 = 1.e-3
            # w2 = 1.e-3
            w3 = 5.e-6

            outputs = torch.stack(predictions, dim=0).to(self.device)
            nll = w3 * torch.nn.functional.nll_loss(outputs.mean(0), target, reduction='sum').to(device=self.device)
            return nll

class KDLoss(nn.Module):
    def __init__(self, temp_factor):
        super(KDLoss, self).__init__()
        self.temp_factor = temp_factor
        self.kl_div = nn.KLDivLoss(reduction="sum")

    def forward(self, input, target):
        log_p = torch.log_softmax(input/self.temp_factor, dim=1)
        q = torch.softmax(target/self.temp_factor, dim=1)
        loss = self.kl_div(log_p, q)*(self.temp_factor**2)/input.size(0)
        return loss

class FocalLoss(nn.Module):
    def __init__(self, gamma=0, alpha=None, size_average=True):
        super(FocalLoss, self).__init__()
        self.gamma = gamma
        self.alpha = alpha
        # if isinstance(alpha,(float,int,long)): self.alpha = torch.Tensor([alpha,1-alpha])
        if isinstance(alpha, (float, int)): self.alpha = torch.Tensor([alpha, 1 - alpha])
        if isinstance(alpha,list): self.alpha = torch.Tensor(alpha)
        self.size_average = size_average

    def forward(self, input, target):
        """
        :param input: [batch, 2]
        :param target: [batch, 1]
        :return:
        """
        if input.dim()>2:
            input = input.view(input.size(0),input.size(1),-1)  # N,C,H,W => N,C,H*W
            input = input.transpose(1,2)    # N,C,H*W => N,H*W,C
            input = input.contiguous().view(-1,input.size(2))   # N,H*W,C => N*H*W,C
        target = target.view(-1,1)

        logpt = F.log_softmax(input, dim=1)
        logpt = logpt.gather(1,target)
        logpt = logpt.view(-1)
        pt = Variable(logpt.data.exp())

        if self.alpha is not None:
            if self.alpha.type()!=input.data.type():
                self.alpha = self.alpha.type_as(input.data)
            at = self.alpha.gather(0,target.data.view(-1))
            logpt = logpt * Variable(at)

        loss = -1 * (1-pt)**self.gamma * logpt
        if self.size_average: return loss.mean()
        else: return loss.sum()


def get_input_optimizer(input_img):
    optimizer = torch.optim.Adam([input_img.requires_grad_()], lr=0.001)
    return optimizer


class GaussianSmoothing(nn.Module):
    """
    Apply gaussian smoothing on a
    1d, 2d or 3d tensor. Filtering is performed seperately for each channel
    in the input using a depthwise convolution.
    Arguments:
        channels (int, sequence): Number of channels of the input tensors. Output will
            have this number of channels as well.
        kernel_size (int, sequence): Size of the gaussian kernel.
        sigma (float, sequence): Standard deviation of the gaussian kernel.
        dim (int, optional): The number of dimensions of the data.
            Default value is 2 (spatial).
    """
    def __init__(self, channels, kernel_size, sigma, dim=3):
        super(GaussianSmoothing, self).__init__()
        if isinstance(kernel_size, numbers.Number):
            kernel_size = [kernel_size] * dim
        if isinstance(sigma, numbers.Number):
            sigma = [sigma] * dim

        # The gaussian kernel is the product of the
        # gaussian function of each dimension.
        kernel = 1
        meshgrids = torch.meshgrid(
            [
                torch.arange(size, dtype=torch.float32)
                for size in kernel_size
            ]
        )
        for size, std, mgrid in zip(kernel_size, sigma, meshgrids):
            mean = (size - 1) / 2
            kernel *= 1 / (std * math.sqrt(2 * math.pi)) * \
                      torch.exp(-((mgrid - mean) / std) ** 2 / 2)

        # Make sure sum of values in gaussian kernel equals 1.
        kernel = kernel / torch.sum(kernel)

        # Reshape to depthwise convolutional weight
        kernel = kernel.view(1, 1, *kernel.size())
        kernel = kernel.repeat(channels, *[1] * (kernel.dim() - 1))
        kernel = kernel.cuda()
        self.register_buffer('weight', kernel)
        self.groups = channels
        self.kernel_size = kernel_size
        if dim == 1:
            self.conv = F.conv1d
        elif dim == 2:
            self.conv = F.conv2d
        elif dim == 3:
            self.conv = F.conv3d
        else:
            raise RuntimeError(
                'Only 1, 2 and 3 dimensions are supported. Received {}.'.format(dim)
            )

    def forward(self, input):
        """
        Apply gaussian filter to input.
        Arguments:
            input (torch.Tensor): Input to apply gaussian filter on.
        Returns:
            filtered (torch.Tensor): Filtered output.
        """
        input = nn.ConstantPad3d(self.kernel_size[0]//2, 0)(input)
        return self.conv(input, weight=self.weight, groups=self.groups)

def model_save_through_validation(fold, epoch, EMS, selected_EMS, ES, model, dir_save_model, metric_1 = 'val_loss', metric_2=None, save_flag = False):
    """ save the model """
    start_eval_epoch = st.early_stopping_start_epoch
    # start_eval_epoch = 1

    tmp_flag = False
    if save_flag == False:
        if epoch >= start_eval_epoch:
            if metric_1 == 'val_loss' or metric_1 == 'val_mean_loss':
                ES(EMS.dict_val_metric[metric_1][-1], None)
                if ES.early_stop == False:
                    # loss
                    if selected_EMS.dict_val_metric[metric_1] >= EMS.dict_val_metric[metric_1][-1]:
                        selected_EMS.selected_ep = epoch
                        selected_EMS.dict_val_metric[metric_1] = EMS.dict_val_metric[metric_1][-1]

                        """save model"""
                        if selected_EMS.latest_selceted_model_dir != '':
                            os.remove(selected_EMS.latest_selceted_model_dir)
                        current_model_dir = '%s/fold%d_epoch%d.ckpt' % (dir_save_model, (fold), (epoch))
                        try:
                            torch.save(model.state_dict(), current_model_dir)
                        except KeyboardInterrupt:
                            pass
                        except ValueError:
                            pass
                        selected_EMS.latest_selceted_model_dir = current_model_dir
                        tmp_flag = True
            else:
                ES(None, EMS.dict_val_metric[metric_1][-1])
                if ES.early_stop == False:
                    # accuracy, AUC
                    if selected_EMS.dict_val_metric[metric_1] <= EMS.dict_val_metric[metric_1][-1]:
                        selected_EMS.selected_ep = epoch
                        selected_EMS.dict_val_metric[metric_1] = EMS.dict_val_metric[metric_1][-1]

                        """save model"""
                        if selected_EMS.latest_selceted_model_dir != '':
                            os.remove(selected_EMS.latest_selceted_model_dir)
                        current_model_dir = '%s/fold%d_epoch%d.ckpt' % (dir_save_model, (fold), (epoch))
                        try:
                            torch.save(model.state_dict(), current_model_dir)
                        except KeyboardInterrupt:
                            pass
                        except ValueError:
                            pass
                        selected_EMS.latest_selceted_model_dir = current_model_dir
                        tmp_flag = True
            print('')
            print('------ metric_{} ------'.format(metric_1))
            print('Selected_epoch : {}'.format(selected_EMS.selected_ep))
            print('Selected_val_metric : {}'.format(selected_EMS.dict_val_metric[metric_1]))
            print('')

        else:
            if selected_EMS.latest_selceted_model_dir != '':
                os.remove(selected_EMS.latest_selceted_model_dir)

            current_model_dir = '%s/fold%d_epoch%d.ckpt' % (dir_save_model, (fold), (epoch))
            try:
                torch.save(model.state_dict(), current_model_dir)
            except KeyboardInterrupt:
                pass
            except ValueError:
                pass
            selected_EMS.latest_selceted_model_dir = current_model_dir
            tmp_flag = True
            print('')
            print('------ metric_{} ------'.format(metric_1))
            print('Selected_epoch : {}'.format(selected_EMS.selected_ep))
            print('Selected_val_metric : {}'.format(selected_EMS.dict_val_metric[metric_1]))
            print('')
    else:
        """save model"""
        if selected_EMS.latest_selceted_model_dir_2 != '':
            os.remove(selected_EMS.latest_selceted_model_dir_2)
        current_model_dir = '%s/fold%d_epoch%d.ckpt' % (dir_save_model, (fold), (epoch))
        try:
            torch.save(model.state_dict(), current_model_dir)
        except KeyboardInterrupt:
            pass
        except ValueError:
            pass
        selected_EMS.latest_selceted_model_dir_2 = current_model_dir
        tmp_flag = True
    return tmp_flag

class eval_selected_metirc_storage():
    def __init__(self):
        super(eval_selected_metirc_storage, self).__init__()

        """ saved model info"""
        self.latest_selceted_model_dir = ''
        self.latest_selceted_model_dir_2 = ''
        self.selected_ep = 0
        self.dict_val_metric = {
            'val_mean_loss': 10000,
            'val_loss': 10000,
            'val_acc': 0,
            'val_auc': 0,
        }

class eval_metric_storage():
    def __init__(self):
        super(eval_metric_storage, self).__init__()

        """ learning rate """
        self.LR = []

        """ train """
        self.total_train_iter = 0
        self.total_train_step = 0
        self.train_loss = []

        self.train_aux_loss= [[] for _ in range(10)]
        self.train_acc = []
        self.train_step = []

        """ val """
        self.dict_val_metric = {
            'val_loss_queue': deque([]),
            'val_mean_loss': [],
            'val_loss': [],
            'val_acc': [],
            'val_acc_aux': [[] for i in range(len(st.lambda_aux))],
            'val_auc': [],
            'val_MAE': [],
            'val_loss_age': [],
        }

        self.val_step = []
        self.val_loss_1 = []
        self.val_loss_2 = []
        self.val_loss_3 = []
        self.val_loss_4 = []
        self.val_loss_5 = []

        """ test """
        self.test_loss = []
        self.test_acc = []
        self.test_acc_aux = [[] for i in range(len(st.lambda_aux))]
        self.test_auc = []
        self.test_MAE = []
        self.test_loss_age = []
        self.test_step = []

        self.test_loss_1 = []
        self.test_loss_2 = []
        self.test_loss_3 = []
        self.test_loss_4 = []
        self.test_loss_5 = []

    def forward(self):
        pass



class EarlyStopping():
    """
    Early Stopping to terminate training early under certain conditions
    """
    def __init__(self, delta=0, patience=5, verbose = True):
        self.delta = delta
        self.patience = patience
        self.verbose = verbose
        self.counter = 0
        self.best_loss = None
        self.best_acc = None
        self.best_mean_loss = None
        self.early_stop = False
        self.val_loss_min = np.Inf
        self.val_acc_max = 0
        self.wait = 0
        self.stopped_epoch = 0
        super(EarlyStopping, self).__init__()

    def __call__(self, val_loss, val_acc):
        if self.early_stop == False:
            if val_loss != None:
                if self.best_loss is None:
                    self.best_loss = val_loss
                # better model has been found.
                if val_loss < self.best_loss + self.delta:
                    self.best_loss = val_loss
                    self.counter = 0
                # saved model is better.
                else:
                    self.counter += 1
                    if self.counter >= self.patience:
                        self.early_stop = True
            else:
                if self.best_acc is None:
                    self.best_acc= val_acc
                # better model has been found.
                if val_acc > self.best_acc + self.delta:
                    self.best_acc = val_acc
                    self.counter = 0
                # saved model is better.
                else:
                    self.counter += 1
                    if self.counter >= self.patience:
                        self.early_stop = True

            if self.verbose == True:
                print(f'Early Stopping counter : {self.counter} out of {self.patience}')
        else:
            pass




def min_max_norm(img):
    max_val = img.max()
    min_val = img.min()
    if max_val==min_val:
        return img
    else:
        norm_img = (img-min_val)/(max_val-min_val)
        return norm_img


def Gauss_Norm_voxelWise(X, mu=0, sigma=1 , train = True):
    # print('Gaussian Normalize')
    if train == True:
        mu = np.expand_dims(np.mean(X, 0), axis=0)  # (1, 176, 210, 168)
        # TODO sigma != 0 , sigma != inf
        sigma = np.expand_dims(np.std(X, 0), axis=0) + np.spacing(1)  # (1, 176, 210, 168)

        norm_Data = np.zeros([X.shape[0],X.shape[1],X.shape[2],X.shape[3]])
        norm_Data[:, :, :, :] = np.nan_to_num((X[:, :, :, :] - mu) / sigma)
        if np.sum(np.isinf(norm_Data[:, :, :, :])) != 0:
            norm_Data[:, :, :, :][np.isinf(norm_Data[:, :, :, :])] = 0
            print("isinf!!!")
            assert np.sum(np.isinf(norm_Data[:, :, :, :])) != 0
        # for i in range(X.shape[0]):
        #     # print('train %d'%i)
        #     norm_Data[i,:,:,:] = np.nan_to_num((X[i,:,:,:]-mu) / sigma)
        #     if np.sum(np.isinf(norm_Data[i,:,:,:])) != 0 :
        #         norm_Data[i,:,:,:][np.isinf(norm_Data[i,:,:,:])] = 0
        #         print("isinf!!!")
        #         assert np.sum(np.isinf(norm_Data[i, :, :, :])) != 0

        return norm_Data, mu, sigma

    else:
        norm_Data = np.zeros([X.shape[0], X.shape[1], X.shape[2], X.shape[3]])
        norm_Data[:, :, :, :] = np.nan_to_num((X[:, :, :, :] - mu) / sigma)
        if np.sum(np.isinf(norm_Data[:, :, :, :])) != 0:
            norm_Data[:,:,:,:][np.isinf(norm_Data[:,:,:,:])] = 0
            print("isinf!!!")
            print(np.sum(np.isinf(norm_Data[:,:,:,:])))
            assert np.sum(np.isinf(norm_Data[:, :, :, :])) != 0
        # for i in range(X.shape[0]):
        #     # print('test %d' % i)
        #     norm_Data[i, :, :, :] = np.nan_to_num((X[i, :, :, :] - mu) / sigma)
        #     # print(np.sum(np.isinf(norm_Data[i,:,:,:])))
        #     if np.sum(np.isinf(norm_Data[i,:,:,:])) != 0:
        #         # norm_Data[i,:,:,:][np.isinf(norm_Data[i,:,:,:])] = 0
        #         print("isinf!!!")
        #         assert np.sum(np.isinf(norm_Data[i, :, :, :])) != 0
        return norm_Data


def Gauss_Norm_subjectWise(X, mu=0, sigma=1 , train = True):
    if train == True:
        mu = np.mean(X,  keepdims=True)# (1,1,1,1)
        sigma = np.std(X, keepdims=True) + np.spacing(1)# (1,1,1,1)

        # TODO sigma != 0 , sigma != inf
        assert sigma != 0
        assert sigma != np.inf
        # norm_Data = np.zeros([X.shape[0],X.shape[1],X.shape[2],X.shape[3]])
        # norm_Data[:, :, :, :] = np.nan_to_num((X[:, :, :, :] - mu) / sigma)
        # norm_Data = np.nan_to_num((X[:, :, :, :] - mu) / sigma)

        norm_Data = (X - mu) / sigma
        assert np.sum(np.isinf(norm_Data)) == 0
        # norm_Data[:,:,:,:][np.isinf(norm_Data[:,:,:,:])] = 0

        return norm_Data, mu, sigma

    else :
        # norm_Data = np.zeros([X.shape[0], X.shape[1], X.shape[2], X.shape[3]])
        # norm_Data[:, :, :, :] = np.nan_to_num((X[:, :, :, :] - mu) / sigma)
        # norm_Data = np.nan_to_num((X[:, :, :, :] - mu) / sigma)
        assert sigma != 0
        assert sigma != np.inf
        norm_Data = (X - mu) / sigma
        assert np.sum(np.isinf(norm_Data)) == 0

        return norm_Data


def data_normalization(X, min = 0 , max = 255):
    norm_Data = np.subtract(X, min)
    norm_Data = np.nan_to_num(np.divide(norm_Data, (max - min)/2))
    norm_Data = np.subtract(norm_Data , 1)
    return norm_Data

def data_minmax(X):
    shape = [None]*len(X)
    for i in range(len(X)):
        shape[i]=X.size()[i]
    for i in range(shape[0]):
        for j in range(shape[1]):
            X[i,j,:,:,:] = min_max_norm(X[i,j,:,:,:])

    return X

def data_mean_centering(X):
    norm_Data = np.zeros([X.shape[0], X.shape[1], X.shape[2], X.shape[3], X.shape[4]])
    for i in range(X.shape[0]):
        mu = np.mean(np.array(X[i]), keepdims=True)
        norm_Data[i] = np.subtract(X[i] , mu)
    return norm_Data


def Cross_validation(num_data, k_fold, Random_seed=0):
    indices = np.random.RandomState(seed=Random_seed).permutation(num_data)
    np.random.shuffle(indices)
    num_idx = num_data // k_fold
    sample_remainder = num_data % k_fold
    list_size_each_fold = []
    tmp = 0
    for i_fold in range(k_fold):
        if sample_remainder > i_fold:
            list_size_each_fold.append(num_idx + 1 + tmp)
        else:
            list_size_each_fold.append(num_idx+ tmp)
        tmp = list_size_each_fold[-1]

    train_idx = []
    test_idx = []
    val_idx = []
    for i_fold in range(k_fold):
        fold_slice = np.split(indices.copy(), list_size_each_fold, axis=0)
        fold_slice.pop(-1)
        if i_fold == k_fold - 1:
            test_idx.append(fold_slice.pop(i_fold % 10))
            val_idx.append(fold_slice.pop(0))
        else:
            test_idx.append(fold_slice.pop(i_fold % 10))
            val_idx.append(fold_slice.pop(i_fold % 10))
        train_idx.append(np.concatenate(fold_slice))
    return train_idx, val_idx, test_idx



def search_in_whole_subdir(file_dir, sub_dir, n_file, n_ext='.nii'):
    """
    :param file_dir: file directory
    :param sub_dir: the directory default = ''
    :param n_file: a list which words that extraction included
    :param n_ext: the type of files (e.g., .gt, .nii)
    :return: file list
    """

    """ make dir to save if not exist """
    if os.path.exists(file_dir + sub_dir) == False:
        os.makedirs(file_dir+sub_dir)

    file_list = [] # the list to reture
    for (path, dir, files) in os.walk(file_dir + sub_dir):
        # print(path)
        for filename in files:
            ext = os.path.splitext(filename)[-1] # 0 : filename, 1 : 확장자
            _file = os.path.splitext(filename)[0]
            if ext == n_ext:
                count_ = 0
                for i in range (len(n_file)):
                     if n_file[i] in _file :
                         count_ += 1
                if count_ == len(n_file) :
                    file_to_save = path + '/' + filename
                    file_list.append(file_to_save)
    # print(len(file_list))
    return file_list

def save_tensor_to_img(tensor, name):
    tmp_array = tensor.data.cpu().numpy()
    fig = plt.figure()
    shape = tmp_array.shape
    # plt.imshow(np.asarray(tmp_array[0, 0, int(shape[2]/2), :, :]))
    plt.imshow(np.asarray(tmp_array[0, 0, :, int(shape[3] / 2), :]))
    # plt.imshow(np.asarray(tmp_array[0, 0, : , :, int(shape[4] / 2)]))
    plt.pcolor
    plt.colorbar()
    fig.savefig(st.save_plot_dir + name + ".png")
    plt.close(fig)


def plot_confusion_matrix(y_true, y_pred, classes,f_dir, f_name, normalize=False, title=None, cmap=plt.cm.Blues):
    """
    This function prints and plots the confusion matrix.
    Normalization can be applied by setting `normalize=True`.
    """
    if not title:
        if normalize:
            title = 'Normalized confusion matrix'
        else:
            title = 'Confusion matrix, without normalization'

    # Compute confusion matrix
    cm = confusion_matrix(y_true, y_pred)

    # Only use the labels that appear in the data
    classes = classes[unique_labels(y_true, y_pred)]
    if normalize:
        cm = cm.astype('float') / cm.sum(axis=1)[:, np.newaxis]
        print("Normalized confusion matrix")
    else:
        print('Confusion matrix, without normalization')

    print(cm)

    plt.rcParams.update({'font.size': 10})
    fig, ax = plt.subplots()
    im = ax.imshow(cm, interpolation='nearest', cmap=cmap)
    ax.figure.colorbar(im, ax=ax)

    # We want to show all ticks...
    ax.set(xticks=np.arange(cm.shape[1]),
           yticks=np.arange(cm.shape[0]),
           # ... and label them with the respective list entries
           xticklabels=classes, yticklabels=classes,
           title=title,
           ylabel='True label',
           xlabel='Predicted label')

    # Rotate the tick labels and set their alignment.
    plt.setp(ax.get_xticklabels(), rotation=45, ha="right",
             rotation_mode="anchor")

    # Loop over data dimensions and create text annotations.
    fmt = '.2f' if normalize else 'd'
    thresh = cm.max() / 2.
    for i in range(cm.shape[0]):
        for j in range(cm.shape[1]):
            ax.text(j, i, format(cm[i, j], fmt),
                    ha="center", va="center",
                    color="white" if cm[i, j] > thresh else "black")
    fig.tight_layout()
    plt.savefig(f_dir + f_name)

def save_tensor_to_img(tensor, save_dir):
    tmp_array = tensor.data.cpu().numpy()
    fig = plt.figure()
    shape = tmp_array.shape
    # plt.imshow(np.asarray(tmp_array[0, 0, int(shape[2]/2), :, :]))
    plt.imshow(np.asarray(tmp_array[0, 0, :, int(shape[3] / 2), :]))
    # plt.imshow(np.asarray(tmp_array[0, 0, : , :, int(shape[4] / 2)]))
    plt.pcolor
    plt.colorbar()
    fig.savefig(save_dir)
    plt.close(fig)

def save_numpy_to_2D_img(img, save_dir, file_name= 'test'):
    make_dir(save_dir)
    tmp_array = img
    fig = plt.figure()
    shape = tmp_array.shape
    # plt.imshow(np.asarray(tmp_array[int(shape[0]/2), :, :]))
    plt.imshow(np.asarray(tmp_array[:, int(shape[1] / 2), :]))
    # plt.imshow(np.asarray(tmp_array[: , :, int(shape[2] / 2)]))
    plt.pcolor
    plt.colorbar()
    fig.savefig(save_dir + file_name)
    plt.close(fig)

def save_featureMap_tensor(tensor, dirToSave = './', name='test'):
    tmp_dir = dirToSave +'/featuremap'
    if os.path.exists(tmp_dir) == False:
        os.makedirs(tmp_dir)
    tmp_array = tensor.data.cpu().numpy()
    f_img = nib.Nifti1Image(tmp_array, np.eye(4))
    nib.save(f_img, os.path.join(tmp_dir + '/'+ name+ '.nii.gz'))

def save_featureMap_numpy(numpy, dirToSave = './', name='test'):
    tmp_dir = dirToSave +'/featuremap'
    if os.path.exists(tmp_dir) == False:
        os.makedirs(tmp_dir)
    f_img = nib.Nifti1Image(numpy, np.eye(4))
    nib.save(f_img, os.path.join(tmp_dir + '/'+ name+ '.nii.gz'))


def preparation_fold_index(config):
    list_trIdx = [] # (# of class, # of fold)
    list_valIdx = []  # (# of class, # of fold)
    list_teIdx = []  # (# of class, # of fold)

    for i in range(len(st.list_class_type)):
        if st.list_data_type[st.data_type_num] == 'Density':
            tmp_shape = \
            np.memmap(filename=st.ADNI_fold_image_path[i], mode="r", dtype=np.float64).reshape(-1, st.num_modality,
                                                                                               st.x_size, st.y_size,
                                                                                               st.z_size).shape[0]
        elif st.list_data_type[st.data_type_num] == 'ADNI_JSY':
            tmp_shape = \
            np.memmap(filename=st.ADNI_fold_image_path[i], mode="r", dtype=np.float32).reshape(-1, st.num_modality,
                                                                                               st.x_size, st.y_size,
                                                                                               st.z_size).shape[0]
        elif 'ADNI_Jacob' in st.list_data_type[st.data_type_num] or 'ADNI_AAL_256' in st.list_data_type[st.data_type_num]:
            tmp_shape = \
            np.memmap(filename=st.ADNI_fold_image_path[i], mode="r", dtype=np.uint8).reshape(-1, st.num_modality,
                                                                                               st.x_size, st.y_size,
                                                                                               st.z_size).shape[0]

        print(tmp_shape)
        tmp_trIdx, tmp_valIdx, tmp_teIdx = utils.Cross_validation(tmp_shape, config.kfold, Random_seed=0)
        list_trIdx.append(tmp_trIdx)
        list_valIdx.append(tmp_valIdx)
        list_teIdx.append(tmp_teIdx)

    """ Check whether all of the index is different """
    for i_class_type in range(len(st.list_class_type)):
        for i_fold in range(config.kfold):

            for j_class_type in range(len(st.list_class_type)):
                for j_fold in range(config.kfold):

                    if i_fold != j_fold  or i_class_type != j_class_type:
                        assert not(np.array_equal(list_trIdx[i_class_type][i_fold], list_trIdx[j_class_type][j_fold]))
                        assert not(np.array_equal(list_valIdx[i_class_type][i_fold], list_valIdx[j_class_type][j_fold]))
                        assert not(np.array_equal(list_teIdx[i_class_type][i_fold], list_teIdx[j_class_type][j_fold]))


    """ save index for each class """
    for i_class_type in range(len(st.list_class_type)):
        with open(st.train_index_dir[i_class_type], 'wb') as fp:
            pickle.dump(list_trIdx[i_class_type], fp)
        with open(st.val_index_dir[i_class_type], 'wb') as fp:
            pickle.dump(list_valIdx[i_class_type], fp)
        with open(st.test_index_dir[i_class_type], 'wb') as fp:
            pickle.dump(list_teIdx[i_class_type], fp)

def plot_list_v1(x, y, title ='None', n_xlabel ='x', n_ylabel ='y', save_dir ='', file_name ='', flag ='minmax', flag_match = False):
    np_save_dir = save_dir + '/np'
    if os.path.exists(save_dir) == False:
        os.makedirs(save_dir)
    if os.path.exists(np_save_dir) == False:
        os.makedirs(np_save_dir)
    x_range = [None] * 2
    y_range = [None] * 2
    margin = 0.05

    fig = plt.figure(figsize=(40, 10 * len(y)))
    fig.suptitle(title, fontsize=50)
    plt.rcParams.update({'font.size': 22})

    ##TODO : x_range
    x_range[0] = min(x)
    x_range[1] = max(x)

    ##TODO : y_range
    if flag_match == True:
        if flag == 'minmax':
            y_range[0] = np.vstack(y).min()
            y_range[1] = np.vstack(y).max()

        elif flag == 'percentile':
            y_range[0] = np.percentile(np.vstack(y), 1)
            y_range[1] = np.percentile(np.vstack(y), 99)

        elif flag == 'dist':
            mean = np.vstack(y).mean()
            std = np.vstack(y).std()
            y_range[0] = mean - 5 * std
            y_range[1] = mean + 5 * std

    ##TODO: plotting ans save
    for i in range(len(y)):
        ax1 = fig.add_subplot(len(y), 1, i + 1)
        # ax1.set_title(title + '_{}'.format(i))
        ax1.set_ylabel(n_ylabel[i], color='b')
        ax1.set_xlabel(n_xlabel, color='b')
        ax1.plot(x, y[i], c='b', ls='-', marker='.', label=n_ylabel[i])  # ls : :, -, o-, .-
        plt.grid(True)
        plt.legend()
        if flag_match != True:
            if flag == 'minmax':
                y_range[0] = np.array(y[i]).min()
                y_range[1] = np.array(y[i]).max()

            elif flag == 'percentile':
                y_range[0] = np.percentile(y[i], 1)
                y_range[1] = np.percentile(y[i], 99)

            elif flag == 'dist':
                mean = np.array(y[i]).mean()
                std = np.array(y[i]).std()
                y_range[0] = mean - 5 * std
                y_range[1] = mean + 5 * std
        if (x_range[1] - x_range[0]) > 0:
            plt.xlim(x_range[0] - (x_range[1] - x_range[0]) * margin, x_range[1] + (x_range[1] - x_range[0]) * margin)
        if (y_range[1] - y_range[0]) > 0:
            plt.ylim(y_range[0] - (y_range[1] - y_range[0]) * margin, y_range[1] + (y_range[1] - y_range[0]) * margin)
        np.save(file=np_save_dir + file_name + '_' + n_ylabel[i], arr=y[i])

    plt.savefig(os.path.join(save_dir + file_name))

    plt.close('all')

def plot_list_v2(x, y, title ='None', n_xlabel ='x', n_ylabel ='y', save_dir ='', file_name ='', flag ='minmax'):
    np_save_dir = save_dir + '/np'
    if os.path.exists(save_dir) == False:
        os.makedirs(save_dir)
    if os.path.exists(np_save_dir) == False:
        os.makedirs(np_save_dir)

    x_range = [None] * 2
    y_range = [None] * 2
    margin = 0.05

    fig = plt.figure(figsize=(40, 10))
    fig.suptitle(title, fontsize=50)
    plt.rcParams.update({'font.size': 22})
    ##TODO : x_range
    x_range[0] = min(x)
    x_range[1] = max(x)

    ##TODO : y_range
    if flag == 'minmax':
        y_range[0] = np.vstack(y).min()
        y_range[1] = np.vstack(y).max()

    elif flag == 'percentile':
        y_range[0] = np.percentile(np.vstack(y), 1)
        y_range[1] = np.percentile(np.vstack(y), 99)

    elif flag == 'dist':
        mean = np.vstack(y).mean()
        std = np.vstack(y).std()
        y_range[0] = mean - 5 * std
        y_range[1] = mean + 5 * std

    ##TODO: plotting ans save
    ax1 = fig.add_subplot(1, 1, 1)
    # ax1.set_title(title)
    # ax1.set_ylabel(n_ylabel, color='k')
    ax1.set_xlabel(n_xlabel, color='k')

    list_color = ['b', 'g', 'r', 'c', 'm', 'y', 'k']
    for i in range(len(y)):
        ax1.plot(x, y[i], c=list_color[i], ls='-', marker='.', label=n_ylabel[i])  # ls : :, -, o-, .-
        np.save(file=np_save_dir + file_name +'_' + n_ylabel[i], arr=y[i])

    if (x_range[1] - x_range[0]) > 0:
        plt.xlim(x_range[0] - (x_range[1] - x_range[0]) * margin, x_range[1] + (x_range[1] - x_range[0]) * margin)
    if (y_range[1] - y_range[0]) > 0:
        plt.ylim(y_range[0] - (y_range[1] - y_range[0]) * margin, y_range[1] + (y_range[1] - y_range[0]) * margin)

    plt.grid(True)
    plt.legend()
    plt.savefig(os.path.join(save_dir + file_name))
    plt.close('all')


def model_freeze(model, requires_grad = False):
    for param in model.parameters():
        param.requires_grad = requires_grad
    # for name, child in model.named_children():
    #     for param in child.parameters():
    #         param.requires_grad = requires_grad
    #     dfs_freeze(child, requires_grad)

def dfs_freeze(model, requires_grad = False):
    # for param in model.parameters():
    #     param.requires_grad = requires_grad
    for name, child in model.named_children():
        for param in child.parameters():
            param.requires_grad = requires_grad
        dfs_freeze(child, requires_grad)

def model_dir_to_load(fold, model_load_dir):
    """ find the maximum epoch model between saved models"""
    included_file_name = ['fold' + str(fold)]
    # get the model corresponding to the specific fold
    models = search_in_whole_subdir('', model_load_dir, included_file_name, '.ckpt')
    s_index = 0 # start index
    e_index = 0 # end index
    n_epoch = []
    for i in range (len(models)):
        for j in range (len(models[i])):
            if models[i][-(j+1)] == 'h':
                s_index = j
                break

        for j in range (len(models[i])):
            if models[i][-(j+1)] == '.':
                e_index = j
                break
        n_epoch.append(models[i][-(s_index+1)+1 : -(e_index+1)])

    if len(n_epoch) == 0:
        print("There is no selected model!")
        return None
    else:
        included_file_name.append(max(n_epoch))
        # get the model corresponding to the specific max epoch
        models = search_in_whole_subdir('', model_load_dir, included_file_name, '.ckpt')
        model_dir = models[0]
    return model_dir

def model_dir_to_load_2(fold, model_load_dir):
    """ find the maximum epoch model between saved models"""
    included_file_name = ['fold' + str(fold)]
    # get the model corresponding to the specific fold
    models = search_in_whole_subdir('', model_load_dir, included_file_name, '.ckpt')
    s_index = 0 # start index
    e_index = 0 # end index
    n_epoch = []
    for i in range (len(models)):
        for j in range (len(models[i])):
            if models[i][-(j+1)] == 'h':
                s_index = j
                break

        for j in range (len(models[i])):
            if models[i][-(j+1)] == '.':
                e_index = j
                break
        n_epoch.append(models[i][-(s_index+1)+1 : -(e_index+1)])
    included_file_name.append(max(n_epoch))
    # get the model corresponding to the specific max epoch
    models = search_in_whole_subdir('', model_load_dir, included_file_name, '.ckpt')
    model_dir = models[0]
    return model_dir

def tensor_cropping(x, RF_size = 33) :
    """
    x =  tensor
    RF_size = receptive_filed_size
    """
    tmp_size = [None]*3
    for i in range(3):
        tmp_size[i] = x.size()[i+2]

    standard = RF_size // 2
    tmp_coord = np.random.randint(standard, size=3)
    out = x[:, :,
          tmp_coord[0]: tmp_size[0] - standard + tmp_coord[0],
          tmp_coord[1]: tmp_size[1] - standard + tmp_coord[1],
          tmp_coord[2]: tmp_size[2] - standard + tmp_coord[2],
          ]
    return out

def smooth_one_hot(true_labels, classes=2, smoothing=0.9):
    """
    if smoothing == 0, it's one-hot method
    if 0 < smoothing < 1, it's smooth method

    """
    assert 0 <= smoothing < 1
    confidence = 1.0 - smoothing
    label_shape = torch.Size((true_labels.size(0), classes))
    with torch.no_grad():
        true_dist = torch.empty(size=label_shape, device=true_labels.device)
        true_dist.fill_(smoothing / (classes - 1))
        # true_dist.scatter_(1, true_labels.data.unsqueeze(1), confidence)
        true_dist.scatter_(1, true_labels.data, confidence)
    return true_dist

def roll_tensor(x, n, axis):
    if axis == 0 :
        return torch.cat((x[-n:, :, :], x[:-n, :, ]), dim=0)
    elif axis == 1 :
        return torch.cat((x[:, -n:, :], x[:, :-n, ]), dim=1)
    elif axis == 2 :
        return torch.cat((x[:, :, -n:], x[:, :, :-n]), dim=2)

def push_tensor(x, n, axis):
    zero_tensor = torch.zeros_like(x)
    if axis == 0 :
        if n > 0 :
            return torch.cat((zero_tensor[-n:, :, :], x[:-n, :, :]), dim=0)
        elif n < 0 :
            return torch.cat((x[-n:, :, :], zero_tensor[:-n, :, :]), dim=0)
        else:
            return x
    elif axis == 1 :
        if n > 0:
            return torch.cat((zero_tensor[:, -n:, :], x[:, :-n, :]), dim=1)
        elif n < 0:
            return torch.cat((x[:, -n:, :], zero_tensor[:, :-n, :]), dim=1)
        else:
            return x

    elif axis == 2 :
        if n > 0:
            return torch.cat((zero_tensor[:, :, -n:], x[:, :, :-n]), dim=2)
        elif n < 0:
            return torch.cat((x[:, :, -n:], zero_tensor[:, :, :-n]), dim=2)
        else:
            return x


def crop_tensor(datas):
    list_cropping_info = [[] for tmp_i in range(3)]
    """ width """
    tmp_size = [a_i - b_i for a_i, b_i in zip(st.max_crop_size, st.min_crop_size)]
    width_size = [np.random.randint(k) + st.min_crop_size[i] for i, k in enumerate(tmp_size)]

    """ start """
    img_size = [st.x_size , st.y_size, st.z_size]
    start_point = [np.random.randint(img_size[i] - i_width) for i, i_width in enumerate(width_size)]

    for i in range(len(img_size)):
        list_cropping_info[i].append(start_point[i])
        list_cropping_info[i].append(width_size[i])

    return datas[:, :, list_cropping_info[0][0]:list_cropping_info[0][0] + list_cropping_info[0][1],
            list_cropping_info[1][0]:list_cropping_info[1][0] + list_cropping_info[1][1],
            list_cropping_info[2][0]:list_cropping_info[2][0] + list_cropping_info[2][1]], list_cropping_info




def excel_setting(start_fold, end_fold, result_dir, f_name):
    """ setting for the excel file """
    wb = Workbook()
    ws1 = wb.create_sheet('train_result', 0)
    exp_name = st.exp_title
    exp_description = st.exp_description

    """excel setting"""

    """ first col"""
    ws1.cell(row=1 + st.push_start_row, column=1, value="fold")
    for i in range(len(st.list_eval_metric)):
        ws1.cell(row=2+i + st.push_start_row, column=1, value=st.list_eval_metric[i])

    """ first row"""
    for col in range(start_fold, end_fold + 1):
        ws1.cell(row=1 + st.push_start_row, column=col + 1, value="fold_" + str(col))
    ws1.cell(row=1 + st.push_start_row, column=end_fold + 2, value="Avg")
    ws1.cell(row=1 + st.push_start_row, column=end_fold + 2).font = Font(name='Calibri', size=12, bold=True)
    column = str(chr(64 + end_fold + 2))
    ws1.column_dimensions[column].width = 20

    """ head """
    n_row = ws1.max_row
    n_col = ws1.max_column
    ws1.merge_cells(start_row=1, end_row = 1, start_column= 1, end_column=n_col)
    ws1.merge_cells(start_row=2, end_row = 2, start_column= 1, end_column=n_col)
    ws1.cell(row=1, column=1, value=exp_name)
    ws1.cell(row=2, column=1, value=exp_description)


    box = Border(left=Side(style='thin'),
                 right=Side(style='thin'),
                 top=Side(style='thin'),
                 bottom=Side(style='thin'))

    """save xlsx"""
    n_row = ws1.max_row
    n_col = ws1.max_column
    ws1.column_dimensions['A'].width = 20
    for i_row in range(1, n_row+1):
        for i_col in range(1, n_col+1):
            ca1 = ws1.cell(row = i_row, column = i_col)
            ca1.alignment = Alignment(horizontal='center', vertical='center')
            ca1.border = box
            if i_col == 1:
                ca1.font = Font(name='Calibri', size = 15, bold=True)
    wb.save(result_dir + "/{}.xlsx".format(f_name))
    return wb, ws1

def analysis_AttentionMap(fold, attn, batch_num, label, dim ):
    # ut.analysis_AttentionMap(fold=fold, attn=attn, batch_num=test_batchnum, label=labels.data.cpu().numpy().squeeze(), dim=(11, 14, 11))

    tmp = attn[0].data.cpu().numpy()
    tmp_dir = './attention_analysis'
    if os.path.exists(tmp_dir) == False:
        os.makedirs(tmp_dir)

    for i_head in range(tmp.shape[0]):
        std = np.std(tmp[i_head], axis=0)
        mean = np.mean(tmp[i_head], axis=0)
        x = np.arange(0, tmp.shape[-1], 1)

        fig = plt.figure()
        ax = fig.add_subplot(111)
        ax_1 = fig.add_subplot(211)
        ax_2 = fig.add_subplot(212)

        # Turn off axis lines and ticks of the big subplot
        # ax.spines['top'].set_color('none')
        # ax.spines['bottom'].set_color('none')
        # ax.spines['left'].set_color('none')
        # ax.spines['right'].set_color('none')
        ax.tick_params(labelcolor='w', top=False, bottom=False, left=False, right=False)

        ax_1.plot(x, mean)
        ax_2.plot(x, std)

        # Set common labels
        ax.set_xlabel('position')
        # ax.set_ylabel('mean & std')

        ax_1.set_title('mean')
        ax_2.set_title('std')

        # plt.show()
        plt.savefig(tmp_dir + '/sum_std_of_attn_label_{}_batchN_{}_head_{}_fold_{}'.format(label, batch_num, i_head, fold))

        save_featureMap_numpy(mean.reshape(dim), './attention_mean', 'mean_label_{}_batchN_{}_head_{}_fold_{}'.format(label, batch_num, i_head, fold))
        save_featureMap_numpy(std.reshape(dim), './attention_std', 'std_label_{}_batchN_{}_head_{}_fold_{}'.format(label, batch_num, i_head, fold))

        # plt.imshow(tmp[i_head])
        # plt.colorbar()
        # plt.savefig(tmp_dir + '/a_a_matrix_head_{}_fold_{}'.format(i_head, fold))
        plt.close('all')
    return None

def make_dir(dir = './', flag_rm = False, flag_assert = False):
    if flag_rm == True:
        shutil.rmtree(dir)
        os.makedirs(dir)
    else:
        if os.path.exists(dir) == False:
            os.makedirs(dir)
        else :
            if flag_assert == True:
                assert os.path.exists(dir) == False


def copy_dir(src, dst, flag_rm = False, flag_assert = False):
    if flag_rm == True:
        if os.path.exists(dst) == True:
            shutil.rmtree(dst)
        shutil.copytree(src=src, dst=dst)
    else:
        if os.path.exists(dst) == True:
            if flag_assert == True:
                assert os.path.exists(dst) == False, 'The file dir has been already existed!!!!'
        else:
            shutil.copytree(src=src, dst=dst)


def eval_classification_model_bayesian(config, fold, loader, model, criterion, confusion_save_info = None):
    """ loader"""
    correct = 0
    loss_np = 0
    model.eval()
    torch.cuda.empty_cache()
    predict_result_cls = []
    groundTruth_cls = []
    count = 0

    if fst.flag_RoI_template ==True:
        RoI_template = torch.tensor(nib.load(st.RoI_template_dir).get_data().squeeze()[st.x_range[0] : st.x_range[1], st.y_range[0] : st.y_range[1], st.z_range[0]: st.z_range[1]]).float().cuda()

    with torch.no_grad():
        for datas, labels, alabels, mlabel in loader:
            count +=1

            """ input"""
            datas = Variable(datas[:, :, :, :, :].view(-1, 1, config.modality, config.sagital, config.coronal, config.axial)).cuda()
            labels = Variable(labels.long()).cuda()

            """ get the output, logit"""
            model.eval()

            """ forward propagation """
            dict_result = model(datas, sample=False)
            # dict_result = model(datas, alabel.cuda())
            output_logit = dict_result['logits']


            """ calculate the loss """
            predictions = []
            for i in range(config.samples):
                pred = model(datas,sample=False)['logits']
                predictions.append(pred)
            # w3 = 5.e-6
            w3 = 1
            outputs = torch.stack(predictions,dim=0).cuda()
            loss = w3*torch.nn.functional.nll_loss(outputs.mean(0), labels, reduction='sum').cuda()


            """ add the loss """
            loss_np += loss.data.cpu().numpy()

            """ count the correct prediction """
            # prob = nn.Softmax(dim=1)(output_logit)
            _, pred = output_logit.max(1, keepdim=True)
            # pred = prob.argmax(dim=1, keepdim=True)
            correct += pred.eq(labels.view_as(pred)).sum().cpu().item()

            # """ plot on 2D """
            # plot_logit_on_2D(logit_0=output_logitMap[0, 0].cpu().numpy(), logit_1=output_logitMap[0, 1].cpu().numpy(), gt=labels.cpu().item(), pred=pred.cpu().item(),
            #                  save_dir=st.dir_to_save_1 +'/plot_logit_on_2D', save_file_name='sample_{}'.format(count), fold=fold)
            #
            # scatter_AD_logit(AD_logit_img=output_logitMap[0, 1], featureMaps=dict_result['featureMaps'], sigma_factor=2,
            #                  save_dir=st.dir_to_save_1 + '/scatter_AD_logit', save_file_name='sample_{}'.format(count), fold=fold, pred=pred.cpu().item(), label=labels.cpu().item())
            #
            # scatter_AD_logit_with_mask(AD_logit_img=output_logitMap[0, 1], featureMaps=dict_result['featureMaps'], sigma_factor=2,
            #                  save_dir=st.dir_to_save_1 + '/scatter_AD_logit_with_mask', save_file_name='sample_{}'.format(count), fold=fold, pred=pred.cpu().item(), label=labels.cpu().item())

            """ stack ground truth and prediction """
            predict_result_cls.append(pred.data.cpu().numpy().squeeze())
            groundTruth_cls.append(np.hstack(labels.data.cpu().numpy()).squeeze())

    """ stacking """
    groundTruth = np.hstack(groundTruth_cls)
    predict_result = np.hstack(predict_result_cls)

    """ plot the confusion matrix """
    if confusion_save_info != None:
        """ plot and save the confusion matrix """
        class_names = np.zeros(len(st.list_selected_for_train), dtype="U10")
        for i in range (len(st.list_selected_for_train)):
            class_names[i] = st.list_selected_for_train[i]

        ut.plot_confusion_matrix(groundTruth, predict_result, classes=class_names, f_dir=confusion_save_info[1],
                                 f_name='/confusion_fold_' + str(confusion_save_info[0]) + '.png',
                                 title='Confusion matrix, without normalization')

    if len(st.list_selected_for_train) == 2:
        """ Confusion matrix , Accuracy, sensitvity and specificity """
        cm1 = confusion_matrix(groundTruth, predict_result)

        """ calculate the accuracy """
        total1 = sum(sum(cm1))
        accuracy1 = (cm1[0, 0] + cm1[1, 1]) / total1
        sensitivity1 = cm1[0, 0] / (cm1[0, 0] + cm1[0, 1])
        specificity1 = cm1[1, 1] / (cm1[1, 0] + cm1[1, 1])
        fpr, tpr, thresholds = metrics.roc_curve(groundTruth, predict_result, pos_label=1)
        AUC = metrics.auc(fpr, tpr)

    else:
        accuracy1 = correct / len(loader.dataset)
        sensitivity1 = 0
        specificity1 = 0
        AUC = 0

    """ calculate the accuracy and print """
    total = len(loader.dataset)
    loss_np = loss_np / total

    dict_result = {
        "Loss": loss_np,
        "Acc": accuracy1,
        "Sen": sensitivity1,
        "Spe": specificity1,
        "AUC": AUC,
    }
    return dict_result

def eval_classification_model(config, fold, loader, model, criterion_cls, flag_heatmap = False, confusion_save_info = None):
    """ loader"""
    correct = 0
    list_aux_acc = [0] * len(st.lambda_aux)
    loss_np = 0
    model.eval()
    # torch.cuda.empty_cache()
    groundTruth_result = []
    predict_result = []
    count = 0
    count_sample = 0
    count_NC_plot = 0
    count_AD_plot = 0
    if fst.flag_RoI_template ==True:
        RoI_template = torch.tensor(nib.load(st.RoI_template_dir).get_data().squeeze()[st.x_range[0] : st.x_range[1], st.y_range[0] : st.y_range[1], st.z_range[0]: st.z_range[1]]).float().cuda()

    if flag_heatmap == True:
        tmp_save_dir = st.heatmap_dir + '/fold_{}'.format(fold)
        ut.make_dir(dir=tmp_save_dir, flag_rm=False)

    with torch.no_grad():
        for data_batch in loader:
            count +=1

            """ input"""
            datas = Variable(data_batch['data'].float()).cuda()
            # labels = Variable(data_batch['label'].long()).cuda()
            labels = Variable(data_batch['label'].float()).cuda()


            """ minmax norm """
            if st.list_data_norm_type[st.data_norm_type_num] == 'minmax':
                tmp_datas = datas.view(datas.size(0), -1)
                tmp_datas -= tmp_datas.min(1, keepdim=True)[0]
                tmp_datas /= tmp_datas.max(1, keepdim=True)[0]
                datas = tmp_datas.view_as(datas)

            """ get the output, logit"""
            model.eval()

            """ forward propagation """
            if fst.flag_RoI_template == True:
                dict_result = model(datas, RoI_template)
            else:
                dict_result = model(datas)
            # dict_result = model(datas, alabel.cuda())
            output_1 = dict_result['logits']
            output_2 = dict_result['Aux_logits']
            output_3 = dict_result['logitMap']


            loss_list_1 = []
            if fst.flag_loss_1 == True:
                s_labels = ut.smooth_one_hot(labels, config.num_classes, smoothing=st.smoothing_img)
                loss_2 = criterion_cls(output_1, s_labels) * st.lambda_major[0]
                loss_list_1.append(loss_2)

            if fst.flag_loss_2 == True:
                for i_tmp in range(len(output_2)):
                    s_labels = ut.smooth_one_hot(labels, config.num_classes, smoothing=st.smoothing_roi)
                    loss_2 = criterion_cls(output_2[i_tmp], s_labels) * st.lambda_aux[i_tmp]
                    loss_list_1.append(loss_2)

            """ add the loss """
            loss = sum(loss_list_1)
            loss_np += loss.data.cpu().numpy() * output_1.size(0)

            """ count the correct prediction """
            # prob = nn.Softmax(dim=1)(output_logit)
            # pred = prob.argmax(dim=1, keepdim=True)

            """ stack ground truth and prediction """
            groundTruth_result.append(np.hstack(labels.data.cpu().numpy()).squeeze())

            pred = torch.round(output_1)
            correct += pred.eq(labels.view_as(pred)).sum().cpu().item()
            predict_result.append(pred.data.cpu().numpy().squeeze())

            if fst.flag_loss_2 == True:
                for tmp_i in range(len(output_2)):
                    pred = torch.round(output_2[tmp_i])
                    list_aux_acc[tmp_i] += pred.eq(labels.view_as(pred)).sum().cpu().item()


            # if flag_heatmap == True:
            #     for i in range(labels.size(0)):
            #         heatmap = torch.zeros_like(RoI_template)
            #         for j in range(90):
            #             heatmap = torch.where(RoI_template == j + 1, output_3[i, 0, j], heatmap)
            #         thresh = 0.1
            #         percentile = 0.01
            #
            #         ut.plot_heatmap_with_overlay(orig_img=datas.data.cpu().numpy()[i, 0], heatmap_img=heatmap.data.cpu().numpy(),
            #                                      save_dir=tmp_save_dir + '/1_binary_{}_{}_per_{}_th_{}_n_{}.png'.format(
            #                                          st.list_selected_for_train[0], st.list_selected_for_train[1],
            #                                          percentile, thresh, count_2),
            #                                      fig_title='GT : {}, logit : {:.4f}, mean :{:.4f}, min :{:.4f}, max:{:.4f}'.format(
            #                                          labels[i], output_1[i], output_3[i].mean(), output_3[0].min(),
            #                                          output_3[0].max()), thresh=thresh, percentile=percentile)
            #
            #         count_2+=1

            # if flag_heatmap == True:
            #     upsampled_logitMap_1 = nn.Upsample(size=datas[0, 0].shape, mode ='trilinear')(output_3[:, 0].unsqueeze(1))
            #     for i in range(labels.size(0)):
            #         thresh = 0.1
            #         percentile = 0.01
            #
            #         ut.plot_heatmap_with_overlay(orig_img=datas.data.cpu().numpy()[i, 0], heatmap_img=upsampled_logitMap_1.data.cpu().numpy()[i, 0],
            #                                      save_dir=tmp_save_dir + '/1_binary_{}_{}_per_{}_th_{}_n_{}.png'.format(
            #                                          st.list_selected_for_train[0], st.list_selected_for_train[1],
            #                                          percentile, thresh, count_2),
            #                                      fig_title='GT : {}, logit : {:.4f}, mean :{:.4f}, min :{:.4f}, max:{:.4f}'.format(
            #                                          labels[i], output_1[i], output_3[i].mean(), output_3[0].min(),
            #                                          output_3[0].max()), thresh=thresh, percentile=percentile)
            #
            #         count_2+=1


            # if flag_heatmap == True:
            #     if count_NC_plot <5 or count_AD_plot<5:
            #         print('count_NC : {}'.format(count_NC_plot))
            #         print('count_AD : {}'.format(count_AD_plot))
            #         for i in range(labels.size(0)):
            #             count_sample += 1
            #             plot_flag = False
            #             if (labels[i] == 0 and count_NC_plot < 5):
            #                 count_NC_plot += 1
            #                 plot_flag = True
            #             elif (labels[i] == 1 and count_AD_plot < 5):
            #                 count_AD_plot += 1
            #                 plot_flag = True
            #
            #             if plot_flag == True:
            #                 for j in range(len(output_3)):
            #                     # upsampled_logitMap_1 = nn.Upsample(size=datas[i, 0].shape, mode='trilinear')(output_3[j][i].unsqueeze(1))
            #                     upsampled_logitMap_1 = F.interpolate(input=output_3[j][i].unsqueeze(1), size=datas[i, 0].shape, mode='trilinear', align_corners=False)
            #
            #                     thresh = 0.0
            #                     percentile = 0.01
            #                     ut.make_dir(tmp_save_dir + '/sample_{}'.format(count_sample))
            #                     ut.plot_heatmap_with_overlay(orig_img=datas.data.cpu().numpy()[i, 0], heatmap_img=upsampled_logitMap_1.data.cpu().numpy()[0, 0],
            #                                                  save_dir=tmp_save_dir + '/sample_{}'.format(count_sample) + '/1_binary_{}_{}_per_{}_th_{}_scale_{}.png'.format(
            #                                                      st.list_selected_for_train[0], st.list_selected_for_train[1],
            #                                                      percentile, thresh, j),
            #                                                  fig_title='GT : {}, logit : {:.4f}, mean :{:.4f}, min :{:.4f}, max:{:.4f}'.format(
            #                                                      labels[i], output_1[i], output_3[j][i].mean(), output_3[j][i].min(),
            #                                                      output_3[j][i].max()), thresh=thresh, percentile=percentile)
            #
            #                 for j in range(len(output_4)):
            #                     # upsampled_logitMap_1 = nn.Upsample(size=datas[i, 0].shape, mode='trilinear')(output_4[j][i].unsqueeze(1))
            #                     upsampled_logitMap_1 = F.interpolate(input=output_4[j][i].unsqueeze(1), size=datas[i, 0].shape, mode='trilinear', align_corners=False)
            #                     thresh = 0.0
            #                     percentile = 0.01
            #                     ut.make_dir(tmp_save_dir + '/sample_{}'.format(count_sample))
            #                     ut.plot_heatmap_with_overlay(orig_img=datas.data.cpu().numpy()[i, 0], heatmap_img=upsampled_logitMap_1.data.cpu().numpy()[0, 0],
            #                                                  save_dir=tmp_save_dir + '/sample_{}'.format(count_sample) + '/1_gate_{}_{}_per_{}_th_{}_scale_{}.png'.format(
            #                                                      st.list_selected_for_train[0], st.list_selected_for_train[1],
            #                                                      percentile, thresh, j),
            #                                                  fig_title='GT : {}, logit : {:.4f}, mean :{:.4f}, min :{:.4f}, max:{:.4f}'.format(
            #                                                      labels[i], output_1[i], output_3[j][i].mean(), output_3[j][i].min(),
            #                                                      output_3[j][i].max()), thresh=thresh, percentile=percentile)



            # """ plot on 2D """
            # plot_logit_on_2D(logit_0=output_logitMap[0, 0].cpu().numpy(), logit_1=output_logitMap[0, 1].cpu().numpy(), gt=labels.cpu().item(), pred=pred.cpu().item(),
            #                  save_dir=st.dir_to_save_1 +'/plot_logit_on_2D', save_file_name='sample_{}'.format(count), fold=fold)
            #
            # scatter_AD_logit(AD_logit_img=output_logitMap[0, 1], featureMaps=dict_result['featureMaps'], sigma_factor=2,
            #                  save_dir=st.dir_to_save_1 + '/scatter_AD_logit', save_file_name='sample_{}'.format(count), fold=fold, pred=pred.cpu().item(), label=labels.cpu().item())
            #
            # scatter_AD_logit_with_mask(AD_logit_img=output_logitMap[0, 1], featureMaps=dict_result['featureMaps'], sigma_factor=2,
            #                  save_dir=st.dir_to_save_1 + '/scatter_AD_logit_with_mask', save_file_name='sample_{}'.format(count), fold=fold, pred=pred.cpu().item(), label=labels.cpu().item())



    """ stacking """
    groundTruth_result = np.hstack(groundTruth_result)
    predict_result = np.hstack(predict_result)

    """ plot the confusion matrix """
    if confusion_save_info != None:
        """ plot and save the confusion matrix """
        class_names = np.zeros(len(st.list_selected_for_train), dtype="U10")
        for i in range (len(st.list_selected_for_train)):
            class_names[i] = st.list_selected_for_train[i]

        ut.plot_confusion_matrix(groundTruth_result, predict_result, classes=class_names, f_dir=confusion_save_info[1],
                                 f_name='/confusion_fold_' + str(confusion_save_info[0]) + '.png',
                                 title='Confusion matrix, without normalization')

    if len(st.list_selected_for_test) == 2:
        """ Confusion matrix , Accuracy, sensitvity and specificity """
        cm1 = confusion_matrix(groundTruth_result, predict_result)

        """ calculate the accuracy """
        total1 = sum(sum(cm1))
        accuracy1 = (cm1[0, 0] + cm1[1, 1]) / total1
        sensitivity1 = cm1[0, 0] / (cm1[0, 0] + cm1[0, 1])
        specificity1 = cm1[1, 1] / (cm1[1, 0] + cm1[1, 1])
        fpr, tpr, thresholds = metrics.roc_curve(groundTruth_result, predict_result, pos_label=1)
        AUC = metrics.auc(fpr, tpr)

    else:
        accuracy1 = correct / len(loader.dataset)
        sensitivity1 = 0
        specificity1 = 0
        AUC = 0

    # if fst.flag_loss_2 == True:
    #     for tmp_i in range(len(st.lambda_aux)):
    #         list_aux_acc[tmp_i] /=len(loader.dataset)

    """ calculate the accuracy and print """
    total = len(loader.dataset)
    loss_np = loss_np / total

    dict_result = {
        "Loss": loss_np,
        "Acc": accuracy1,
        "Acc_aux": list_aux_acc,
        "Sen": sensitivity1,
        "Spe": specificity1,
        "AUC": AUC,
    }
    return dict_result


def eval_classification_model_2(config, fold, loader, model_1, model_2, criterion, confusion_save_info = None):
    """ loader"""
    correct = 0
    loss_np = 0
    model_1.eval()
    model_2.eval()
    torch.cuda.empty_cache()
    predict_result_cls = []
    groundTruth_cls = []
    count = 0
    count_2 = 0
    with torch.no_grad():
        for data_batch in loader:
            count +=1

            """ input"""
            datas = Variable(data_batch['data'].float()).cuda()
            # labels = Variable(data_batch['label'].long()).cuda()
            labels = Variable(data_batch['label'].float()).cuda()


            """ minmax norm """
            if st.list_data_norm_type[st.data_norm_type_num] == 'minmax':
                tmp_datas = datas.view(datas.size(0), -1)
                tmp_datas -= tmp_datas.min(1, keepdim=True)[0]
                tmp_datas /= tmp_datas.max(1, keepdim=True)[0]
                datas = tmp_datas.view_as(datas)

            """ forward propagation """
            dict_result = model_2(datas)
            output_3 = dict_result['logitMap']
            dict_result = model_1(output_3)
            output_1 = dict_result['logits']
            output_3 = dict_result['logitMap']
            loss = criterion(output_1, labels)


            """ add the loss """
            loss_np += loss.data.cpu().numpy() * output_1.size(0)


            """ count the correct prediction """
            # prob = nn.Softmax(dim=1)(output_logit)
            # pred = prob.argmax(dim=1, keepdim=True)
            # correct += pred.eq(labels.view_as(pred)).sum().cpu().item()

            pred = torch.round(output_1)
            correct += pred.eq(labels.view_as(pred)).sum().cpu().item()

            """ stack ground truth and prediction """
            predict_result_cls.append(pred.data.cpu().numpy().squeeze())
            groundTruth_cls.append(np.hstack(labels.data.cpu().numpy()).squeeze())

    """ stacking """
    groundTruth = np.hstack(groundTruth_cls)
    predict_result = np.hstack(predict_result_cls)



    """ plot the confusion matrix """
    if confusion_save_info != None:
        """ plot and save the confusion matrix """
        class_names = np.zeros(len(st.list_selected_for_train), dtype="U10")
        for i in range (len(st.list_selected_for_train)):
            class_names[i] = st.list_selected_for_train[i]

        ut.plot_confusion_matrix(groundTruth, predict_result, classes=class_names, f_dir=confusion_save_info[1],
                                 f_name='/confusion_fold_' + str(confusion_save_info[0]) + '.png',
                                 title='Confusion matrix, without normalization')

    if len(st.list_selected_for_train) == 2:
        """ Confusion matrix , Accuracy, sensitvity and specificity """
        cm1 = confusion_matrix(groundTruth, predict_result)

        """ calculate the accuracy """
        total1 = sum(sum(cm1))
        accuracy1 = (cm1[0, 0] + cm1[1, 1]) / total1
        sensitivity1 = cm1[0, 0] / (cm1[0, 0] + cm1[0, 1])
        specificity1 = cm1[1, 1] / (cm1[1, 0] + cm1[1, 1])
        fpr, tpr, thresholds = metrics.roc_curve(groundTruth, predict_result, pos_label=1)
        AUC = metrics.auc(fpr, tpr)

    else:
        accuracy1 = correct / len(loader.dataset)
        sensitivity1 = 0
        specificity1 = 0
        AUC = 0

    """ calculate the accuracy and print """
    total = len(loader.dataset)
    loss_np = loss_np / total

    dict_result = {
        "Loss": loss_np,
        "Acc": accuracy1,
        "Sen": sensitivity1,
        "Spe": specificity1,
        "AUC": AUC,
    }
    return dict_result


def eval_classification_model_info(config, fold, loader, model, criterion, confusion_save_info = None):
    """ loader"""
    correct = 0
    loss_np = 0
    model.eval()
    torch.cuda.empty_cache()
    predict_result_cls = []
    groundTruth_cls = []
    count = 0
    list_logitMap = []
    if fst.flag_RoI_template ==True:
        RoI_template = torch.tensor(nib.load(st.RoI_template_dir).get_data().squeeze()[st.x_range[0] : st.x_range[1], st.y_range[0] : st.y_range[1], st.z_range[0]: st.z_range[1]]).float().cuda()

    with torch.no_grad():
        for datas, labels, alabels, mlabel in loader:
            count +=1

            """ input"""
            datas = Variable(datas).cuda()
            labels = Variable(labels.long()).cuda()

            """ get the output, logit"""
            model.eval()

            """ forward propagation """
            if fst.flag_RoI_template == True:
                dict_result = model(datas, RoI_template)
            else:
                dict_result = model(datas)
            # dict_result = model(datas, alabel.cuda())

            output_logit = dict_result['logits']
            output_logitMap = dict_result['final_evidence'] # batch, 2, 21, 26, 19
            output_featureMap = dict_result['featureMaps']

            """ calculate the loss """
            loss = criterion(output_logit, labels)
            loss_np += loss.data.cpu().numpy()

            """ count the correct prediction """
            prob = nn.Softmax(dim=1)(output_logit)
            pred = prob.argmax(dim=1, keepdim=True)
            correct += pred.eq(labels.view_as(pred)).sum().cpu().item()

            ##TODO: plotting subject-wise histogram
            """ count fold label prediction logit map"""
            plot_hist_patch_level_AD_logit(output_logitMap, count, fold, labels, pred, save_dir= './AD_logit_dist/fold_{}'.format(fold), file_name='/sample{}'.format(count))

            # """ plot on 2D """
            # plot_logit_on_2D(logit_0=output_logitMap[0, 0].cpu().numpy(), logit_1=output_logitMap[0, 1].cpu().numpy(), gt=labels.cpu().item(), pred=pred.cpu().item(),
            #                  save_dir=st.dir_to_save_1 +'/plot_logit_on_2D', save_file_name='sample_{}'.format(count), fold=fold)
            #
            # scatter_AD_logit(AD_logit_img=output_logitMap[0, 1], featureMaps=dict_result['featureMaps'], sigma_factor=2,
            #                  save_dir=st.dir_to_save_1 + '/scatter_AD_logit', save_file_name='sample_{}'.format(count), fold=fold, pred=pred.cpu().item(), label=labels.cpu().item())
            #
            # scatter_AD_logit_with_mask(AD_logit_img=output_logitMap[0, 1], featureMaps=dict_result['featureMaps'], sigma_factor=2,
            #                  save_dir=st.dir_to_save_1 + '/scatter_AD_logit_with_mask', save_file_name='sample_{}'.format(count), fold=fold, pred=pred.cpu().item(), label=labels.cpu().item())

            """ stack ground truth and prediction """
            list_logitMap.append(output_logitMap[:, 1].data.cpu().numpy())
            predict_result_cls.append(pred.data.cpu().numpy().squeeze())
            groundTruth_cls.append(np.hstack(labels.data.cpu().numpy()).squeeze())

    """ stacking """
    groundTruth = np.hstack(groundTruth_cls)
    predict_result = np.hstack(predict_result_cls)
    logitMap = np.vstack(list_logitMap)

    ##TODO : t_test
    tmp_save_dir = './t_test/fold_{}'.format(fold)
    ut.make_dir(dir=tmp_save_dir, flag_rm=False)
    statistic, pvalue = stats.ttest_ind(logitMap[groundTruth == 0],
                                        logitMap[groundTruth == 1], axis=0, equal_var=False)
    list_thresh = [0.01, 0.001, 0.0001, 0.00001, 0.000001]
    thresh = 0.5
    percentile = 0.1
    sample_orig_img = output_featureMap.mean(dim=1).squeeze().data.cpu().numpy()
    for i, value in enumerate(list_thresh):
        pvalue_1 = 1 * (pvalue < value)
        pvalue_2 = np.where(np.isnan(pvalue_1), 0, pvalue_1)
        pvalue_2 = pvalue_2.astype(np.float64)
        statistic_1 = statistic * (pvalue < value)
        statistic_2 = np.where(np.isnan(statistic_1), 0, statistic_1)
        ut.save_featureMap_numpy(pvalue_2, tmp_save_dir, 'binary_{}'.format(list_thresh[i]))
        ut.save_featureMap_numpy(statistic_2, tmp_save_dir, 'p_value_{}'.format(list_thresh[i]))
        ut.plot_heatmap_with_overlay(orig_img=sample_orig_img, heatmap_img=pvalue_2,
                                     save_dir=tmp_save_dir + '/binary_{}_{}_per_{}_th_{}_n_{}.png'.format(
                                         st.list_selected_for_train[0], st.list_selected_for_train[1], percentile,
                                         thresh, i),
                                     fig_title=str(value),
                                     thresh=thresh,
                                     percentile=percentile)

    #TODO : mean , std
    mean = np.mean(logitMap, axis=0)
    std = np.std(logitMap, axis=0)
    tmp_save_dir = './mean_std/fold_{}'.format(fold)
    ut.save_featureMap_numpy(mean, tmp_save_dir, 'mean_{}_{}'.format(st.list_selected_for_train[0], st.list_selected_for_train[1]))
    ut.save_featureMap_numpy(std, tmp_save_dir,
                             'std_{}_{}'.format(st.list_selected_for_train[0], st.list_selected_for_train[1]))

    thresh = 0.2
    percentile = 0.01
    ut.plot_heatmap_with_overlay(orig_img=sample_orig_img, heatmap_img=mean,
                                 save_dir=tmp_save_dir + '/mean_{}_{}_per_{}_th_{}.png'.format(
                                     st.list_selected_for_train[0], st.list_selected_for_train[1], percentile,
                                     thresh),
                                 fig_title='mean_{}_{}'.format(st.list_selected_for_train[0], st.list_selected_for_train[1]),
                                 thresh=thresh,
                                 percentile=percentile)
    ut.plot_heatmap_with_overlay(orig_img=sample_orig_img, heatmap_img=std,
                                 save_dir=tmp_save_dir + '/std_{}_{}_per_{}_th_{}.png'.format(
                                     st.list_selected_for_train[0], st.list_selected_for_train[1], percentile,
                                     thresh),
                                 fig_title='std_{}_{}'.format(st.list_selected_for_train[0], st.list_selected_for_train[1]),
                                 thresh=thresh,
                                 percentile=percentile)
    """ plot the confusion matrix """
    if confusion_save_info != None:
        """ plot and save the confusion matrix """
        class_names = np.zeros(len(st.list_selected_for_train), dtype="U10")
        for i in range (len(st.list_selected_for_train)):
            class_names[i] = st.list_selected_for_train[i]

        ut.plot_confusion_matrix(groundTruth, predict_result, classes=class_names, f_dir=confusion_save_info[1],
                                 f_name='/confusion_fold_' + str(confusion_save_info[0]) + '.png',
                                 title='Confusion matrix, without normalization')

    if len(st.list_selected_for_train) == 2:
        """ Confusion matrix , Accuracy, sensitvity and specificity """
        cm1 = confusion_matrix(groundTruth, predict_result)

        """ calculate the accuracy """
        total1 = sum(sum(cm1))
        accuracy1 = (cm1[0, 0] + cm1[1, 1]) / total1
        sensitivity1 = cm1[0, 0] / (cm1[0, 0] + cm1[0, 1])
        specificity1 = cm1[1, 1] / (cm1[1, 0] + cm1[1, 1])
        fpr, tpr, thresholds = metrics.roc_curve(groundTruth, predict_result, pos_label=1)
        AUC = metrics.auc(fpr, tpr)

    else:
        accuracy1 = correct / len(loader.dataset)
        sensitivity1 = 0
        specificity1 = 0
        AUC = 0

    """ calculate the accuracy and print """
    total = len(loader.dataset)
    loss_np = loss_np / total

    dict_result = {
        "Loss": loss_np,
        "Acc": accuracy1,
        "Sen": sensitivity1,
        "Spe": specificity1,
        "AUC": AUC,
    }
    return dict_result

def eval_classification_model_MC_dropout(config, fold, loader, model, criterion, confusion_save_info = None):
    """ loader"""
    correct = 0
    loss_np = 0
    model.eval()
    torch.cuda.empty_cache()
    predict_result_cls = []
    groundTruth_cls = []
    count = 0
    with torch.no_grad():
        for datas, labels, alabels, mlabel in loader:
            count +=1
            """ input"""
            datas = Variable(datas).cuda()
            labels = Variable(labels.long()).cuda()


            """ get the output, logit"""
            model.eval()


            list_output_logit = []
            list_output_logitMap = []
            for i in range(10):
                """ forward propagation """
                dict_result = model(datas, True)
                output_logit = dict_result['logits']
                output_logitMap = dict_result['logitMap']

                list_output_logit.append(output_logit)
                list_output_logitMap.append(output_logitMap)

            # patch_level_prob_dist = nn.Softmax(dim=1)(torch.cat(list_output_logitMap, dim=0))
            # patch_level_entropy_dist = None
            output_logit = sum(list_output_logit)/len(list_output_logit)

            """ calculate the loss """
            loss = criterion(output_logit, labels)

            """ add the loss """
            loss_np += loss.data.cpu().numpy()

            """ count the correct prediction """
            prob = nn.Softmax(dim=1)(output_logit)
            pred = prob.argmax(dim=1, keepdim=True)
            correct += pred.eq(labels.view_as(pred)).sum().cpu().item()

            """ stack ground truth and prediction """
            predict_result_cls.append(pred.data.cpu().numpy().squeeze())
            groundTruth_cls.append(np.hstack(labels.data.cpu().numpy()).squeeze())

    """ stacking """
    groundTruth = np.hstack(groundTruth_cls)
    predict_result = np.hstack(predict_result_cls)

    """ plot the confusion matrix """
    if confusion_save_info != None:
        """ plot and save the confusion matrix """
        class_names = np.zeros(len(st.list_selected_for_train), dtype="U10")
        for i in range (len(st.list_selected_for_train)):
            class_names[i] = st.list_selected_for_train[i]

        ut.plot_confusion_matrix(groundTruth, predict_result, classes=class_names, f_dir=confusion_save_info[1],
                                 f_name='/confusion_fold_' + str(confusion_save_info[0]) + '.png',
                                 title='Confusion matrix, without normalization')

    if len(st.list_selected_for_train) == 2:
        """ Confusion matrix , Accuracy, sensitvity and specificity """
        cm1 = confusion_matrix(groundTruth, predict_result)

        """ calculate the accuracy """
        total1 = sum(sum(cm1))
        accuracy1 = (cm1[0, 0] + cm1[1, 1]) / total1
        sensitivity1 = cm1[0, 0] / (cm1[0, 0] + cm1[0, 1])
        specificity1 = cm1[1, 1] / (cm1[1, 0] + cm1[1, 1])
        fpr, tpr, thresholds = metrics.roc_curve(groundTruth, predict_result, pos_label=1)
        AUC = metrics.auc(fpr, tpr)

    else:
        accuracy1 = correct / len(loader.dataset)
        sensitivity1 = 0
        specificity1 = 0
        AUC = 0

    """ calculate the accuracy and print """
    total = len(loader.dataset)
    loss_np = loss_np / total

    dict_result = {
        "Loss": loss_np,
        "Acc": accuracy1,
        "Sen": sensitivity1,
        "Spe": specificity1,
        "AUC": AUC,
    }
    return dict_result


def eval_classification_model_esemble(config, fold, loader, model, criterion, confusion_save_info = None):
    """ loader"""
    correct = 0
    loss_np = 0
    model.eval()
    torch.cuda.empty_cache()
    predict_result_cls = []
    groundTruth_cls = []
    with torch.no_grad():
        """ get the output, logit"""
        model.eval()

        for datas, labels, alabels, mlabel in loader:
            stride_for_assemble = st.size_translation // 2
            list_logits = []
            list_loss = []
            for i in range(2):
                for j in range(2):
                    for k in range(2):
                        """ input"""
                        datas = Variable(datas[:, :, i * stride_for_assemble:, j * stride_for_assemble:, k * stride_for_assemble:]).cuda()
                        labels = Variable(labels.long()).cuda()

                        """ forward propagation """
                        dict_result = model(datas)
                        # dict_result = model(datas, alabel.cuda())
                        output_logit = dict_result['logits']


                        """ calculate the loss """
                        loss = criterion(output_logit, labels)
                        list_logits.append(output_logit.data.cpu().numpy())
                        list_loss.append(loss.data.cpu().numpy())

            """ add the loss """
            # loss_np += loss.data.cpu().numpy()
            loss_np += np.mean(list_loss)
            f_output_logit = np.mean(np.vstack(list_logits), axis = 0, keepdims=True)

            """ count the correct prediction """
            prob = nn.Softmax(dim=1)(torch.tensor(f_output_logit).cuda())
            pred = prob.argmax(dim=1, keepdim=True)
            correct += pred.eq(labels.view_as(pred)).sum().cpu().item()

            """ stack ground truth and prediction """
            predict_result_cls.append(pred.data.cpu().numpy().squeeze())
            groundTruth_cls.append(np.hstack(labels.data.cpu().numpy()).squeeze())

    """ stacking """
    groundTruth = np.hstack(groundTruth_cls)
    predict_result = np.hstack(predict_result_cls)

    """ plot the confusion matrix """
    if confusion_save_info != None:
        """ plot and save the confusion matrix """
        class_names = np.zeros(len(st.list_selected_for_train), dtype="U10")
        for i in range (len(st.list_selected_for_train)):
            class_names[i] = st.list_selected_for_train[i]

        ut.plot_confusion_matrix(groundTruth, predict_result, classes=class_names, f_dir=confusion_save_info[1],
                                 f_name='/confusion_fold_' + str(confusion_save_info[0]) + '.png',
                                 title='Confusion matrix, without normalization')

    if len(st.list_selected_for_train) == 2:
        """ Confusion matrix , Accuracy, sensitvity and specificity """
        cm1 = confusion_matrix(groundTruth, predict_result)

        """ calculate the accuracy """
        total1 = sum(sum(cm1))
        accuracy1 = (cm1[0, 0] + cm1[1, 1]) / total1
        sensitivity1 = cm1[0, 0] / (cm1[0, 0] + cm1[0, 1])
        specificity1 = cm1[1, 1] / (cm1[1, 0] + cm1[1, 1])
        fpr, tpr, thresholds = metrics.roc_curve(groundTruth, predict_result, pos_label=1)
        AUC = metrics.auc(fpr, tpr)

    else:
        accuracy1 = correct / len(loader.dataset)
        sensitivity1 = 0
        specificity1 = 0
        AUC = 0

    """ calculate the accuracy and print """
    total = len(loader.dataset)
    loss_np = loss_np / total

    dict_result = {
        "Loss": loss_np,
        "Acc": accuracy1,
        "Sen": sensitivity1,
        "Spe": specificity1,
        "AUC": AUC,
    }
    return dict_result

def eval_classification_model_cropped_input(config, fold, loader, model, criterion, confusion_save_info = None):
    """ loader"""
    correct = 0
    loss_np = 0
    model.eval()
    torch.cuda.empty_cache()
    predict_result_cls = []
    groundTruth_cls = []
    tmp_size_x_1 = (st.x_size - st.max_crop_size) // 2
    tmp_size_x_2 = tmp_size_x_1 + st.max_crop_size

    tmp_size_y_1 = (st.y_size - st.max_crop_size) // 2
    tmp_size_y_2 = tmp_size_y_1 + st.max_crop_size

    tmp_size_z_1 = (st.z_size - st.max_crop_size) // 2
    tmp_size_z_2 = tmp_size_z_1 + st.max_crop_size
    with torch.no_grad():
        for datas, labels, alabels, mlabel in loader:
            """ input"""
            datas = Variable(
                datas[:, :, tmp_size_x_1:tmp_size_x_2, tmp_size_y_1:tmp_size_y_2, tmp_size_z_1:tmp_size_z_2]).cuda()
            labels = Variable(labels.long()).cuda()

            """ get the output, logit"""
            model.eval()

            """ forward propagation """

            dict_result = model(datas)
            # dict_result = model(datas, alabel.cuda())
            output_logit = dict_result['logits']

            """ calculate the loss """
            loss = criterion(output_logit, labels)

            """ add the loss """
            loss_np += loss.data.cpu().numpy()

            """ count the correct prediction """
            prob = nn.Softmax(dim=1)(output_logit)
            pred = prob.argmax(dim=1, keepdim=True)
            correct += pred.eq(labels.view_as(pred)).sum().cpu().item()

            """ stack ground truth and prediction """
            predict_result_cls.append(pred.data.cpu().numpy().squeeze())
            groundTruth_cls.append(np.hstack(labels.data.cpu().numpy()).squeeze())

    """ stacking """
    groundTruth = np.hstack(groundTruth_cls)
    predict_result = np.hstack(predict_result_cls)

    """ plot the confusion matrix """
    if confusion_save_info != None:
        """ plot and save the confusion matrix """
        class_names = np.zeros(len(st.list_selected_for_train), dtype="U10")
        for i in range (len(st.list_selected_for_train)):
            class_names[i] = st.list_selected_for_train[i]

        ut.plot_confusion_matrix(groundTruth, predict_result, classes=class_names, f_dir=confusion_save_info[1],
                                 f_name='/confusion_fold_' + str(confusion_save_info[0]) + '.png',
                                 title='Confusion matrix, without normalization')

    if len(st.list_selected_for_train) == 2:
        """ Confusion matrix , Accuracy, sensitvity and specificity """
        cm1 = confusion_matrix(groundTruth, predict_result)

        """ calculate the accuracy """
        total1 = sum(sum(cm1))
        accuracy1 = (cm1[0, 0] + cm1[1, 1]) / total1
        sensitivity1 = cm1[0, 0] / (cm1[0, 0] + cm1[0, 1])
        specificity1 = cm1[1, 1] / (cm1[1, 0] + cm1[1, 1])
        fpr, tpr, thresholds = metrics.roc_curve(groundTruth, predict_result, pos_label=1)
        AUC = metrics.auc(fpr, tpr)

    else:
        accuracy1 = correct / len(loader.dataset)
        sensitivity1 = 0
        specificity1 = 0
        AUC = 0

    """ calculate the accuracy and print """
    total = len(loader.dataset)
    loss_np = loss_np / total

    dict_result = {
        "Loss": loss_np,
        "Acc": accuracy1,
        "Sen": sensitivity1,
        "Spe": specificity1,
        "AUC": AUC,
    }
    return dict_result




def plot_heatmap_with_overlay(orig_img, heatmap_img, save_dir, fig_title = 'Heatmap', thresh=0.5, percentile = 1):
    shape = heatmap_img.shape
    list_interval = []
    for j in range(3):
        tmp_list = []
        for i in np.arange(30, 71, 5):
            tmp_list.append(int(np.percentile(np.arange(0, shape[j]), i)))
        list_interval.append(np.hstack(tmp_list))

    axis_type = ['Sagittal', 'Coronal', 'Axial']

    fig = plt.figure(figsize=(list_interval[0].shape[0] * 2, len(axis_type) * 3))
    plt.rcParams.update({'font.size': 20})
    fig.suptitle(fig_title, fontsize=30)

    heights = [1] * len(axis_type)
    widths = [10] * (list_interval[0].shape[0])
    widths.append(5)
    gs = gridspec.GridSpec(nrows=len(heights),  # row
                           ncols=len(widths),
                           height_ratios=heights,
                           width_ratios=widths,
                           hspace=0.0,
                           wspace=0.0,
                           )

    cmap_orig = plt.get_cmap('Greys')


    # cmap_heatmap = plt.get_cmap('Reds')
    cmap_heatmap = plt.get_cmap('coolwarm')
    # cmap_heatmap = plt.get_cmap('bwr')

    # for orig
    orig_vmax = np.percentile(orig_img, 100 - percentile)
    orig_vmin = np.percentile(orig_img, percentile)
    # print(orig_vmin, orig_vmax)

    # vmax = 30
    # vmin = -30
    vmax = np.percentile(heatmap_img, 100-percentile)
    vmin = np.percentile(heatmap_img, percentile)
    # vmax = heatmap_img.max()
    # vmin = heatmap_img.min()
    # print(heatmap_img.max())
    # print(heatmap_img.min())
    # print(vmin, vmax)

    if np.abs(vmax) > np.abs(vmin):
        vmax = np.abs(vmax)
        vmin = -np.abs(vmax)
    else:
        vmax = np.abs(vmin)
        vmin = -np.abs(vmin)

    thresh_max = vmax * thresh
    thresh_min = vmin * thresh
    # thresh_max = np.percentile(heatmap_img, 97)
    # thresh_min = np.percentile(heatmap_img, 3)
    # print(thresh_min, thresh_max)
    # if np.abs(thresh_max) < np.abs(thresh_min):
    #     thresh_max = np.abs(thresh_max)
    #     thresh_min = -np.abs(thresh_max)
    # else:
    #     thresh_max = np.abs(thresh_min)
    #     thresh_min = -np.abs(thresh_min)

    alpha = 0.8
    axes = []
    for j, q in enumerate(axis_type):
        for i, p in enumerate(list_interval[j]):

            ax1 = fig.add_subplot(gs[j, i])

            if j == 0:
                orig_scattering_img = np.asarray(orig_img[int(p), :, :])
                heatmap_scattering_img = np.asarray(heatmap_img[int(p), :, :])
            elif j == 1:
                orig_scattering_img = np.asarray(orig_img[:, int(p), :])
                heatmap_scattering_img = np.asarray(heatmap_img[:, int(p), :])
            elif j == 2:
                orig_scattering_img = np.asarray(orig_img[:, :, int(p)])
                heatmap_scattering_img = np.asarray(heatmap_img[:, :, int(p)])

            orig_scattering_img = np.rot90(orig_scattering_img)
            heatmap_scattering_img = np.rot90(heatmap_scattering_img)
            heatmap_scattering_img[
                (heatmap_scattering_img < thresh_max) * (heatmap_scattering_img > thresh_min)] = np.nan

            if i == 0:
                # ax1.set_title(axis_type[j])
                ax1.set_ylabel(axis_type[j])
                # plt.ylabel(axis_type[j])
            ax1.imshow(orig_scattering_img, cmap=cmap_orig, vmin=orig_vmin, vmax=orig_vmax)
            # im = ax1.imshow(heatmap_scattering_img, cmap=cmap_heatmap, alpha=alpha, vmin=positive_vmin, vmax=positive_vmax)
            im = ax1.imshow(heatmap_scattering_img, cmap=cmap_heatmap, alpha=alpha, vmin=vmin, vmax=vmax)
            ax1.set_yticks([])
            ax1.set_xticks([])
            ax1.spines['right'].set_visible(False)
            ax1.spines['top'].set_visible(False)
            ax1.spines['bottom'].set_visible(False)
            ax1.spines['left'].set_visible(False)
            axes.append(ax1)
            # ax1.axis('off')
            del orig_scattering_img, heatmap_scattering_img

    # (left, bottom, width, height)
    cax = plt.axes([0.90, 0.1, 0.02, 0.8]) # left, bottom, width, height
    cbar = fig.colorbar(im, ax=axes, extend='both', cax=cax)

    cbar.set_ticks(np.array((vmin, thresh_min, thresh_max, vmax)))
    cbar.set_ticklabels(["%.2f" % (vmin), "%.2f" % (thresh_min), "%.2f" % (thresh_max), "%.2f" % (vmax)])
    # plt.subplots_adjust(bottom=0.1, right=0.6, top=0.9, left=0.5)

    # plt.tight_layout()
    plt.savefig(save_dir, dpi = 300)
    plt.close('all')


def plot_heatmap_without_overlay(heatmap_img, save_dir, fig_title = 'Heatmap', thresh=0.5, percentile = 1):
    shape = heatmap_img.shape
    list_interval = []
    for j in range(3):
        tmp_list = []
        for i in np.arange(30, 71, 5):
            tmp_list.append(int(np.percentile(np.arange(0, shape[j]), i)))
        list_interval.append(np.hstack(tmp_list))

    axis_type = ['Sagittal', 'Coronal', 'Axial']

    fig = plt.figure(figsize=(list_interval[0].shape[0] * 2, len(axis_type) * 3))
    plt.rcParams.update({'font.size': 20})
    fig.suptitle(fig_title, fontsize=30)

    heights = [1] * len(axis_type)
    widths = [10] * (list_interval[0].shape[0])
    widths.append(10)
    gs = gridspec.GridSpec(nrows=len(heights),  # row
                           ncols=len(widths),
                           height_ratios=heights,
                           width_ratios=widths,
                           hspace=0.0,
                           wspace=0.0,
                           )


    # cmap_heatmap = plt.get_cmap('Reds')
    cmap_heatmap = plt.get_cmap('coolwarm')
    # cmap_heatmap = plt.get_cmap('bwr')


    # vmax = np.percentile(heatmap_img, 100-percentile)
    # vmin = np.percentile(heatmap_img, percentile)
    # vmax = heatmap_img.max()
    # vmin = heatmap_img.min()
    vmax = 30
    vmin = -30
    print(heatmap_img.max())
    print(heatmap_img.min())
    print(vmin, vmax)

    if np.abs(vmax) > np.abs(vmin):
        vmax = np.abs(vmax)
        vmin = -np.abs(vmax)
    else:
        vmax = np.abs(vmin)
        vmin = -np.abs(vmin)

    thresh_max = vmax * thresh
    thresh_min = vmin * thresh

    # thresh_max = np.percentile(heatmap_img, 97)
    # thresh_min = np.percentile(heatmap_img, 3)
    # print(thresh_min, thresh_max)
    # if np.abs(thresh_max) < np.abs(thresh_min):
    #     thresh_max = np.abs(thresh_max)
    #     thresh_min = -np.abs(thresh_max)
    # else:
    #     thresh_max = np.abs(thresh_min)
    #     thresh_min = -np.abs(thresh_min)

    alpha = 1
    axes = []
    for j, q in enumerate(axis_type):
        for i, p in enumerate(list_interval[j]):

            ax1 = fig.add_subplot(gs[j, i])

            if j == 0:
                heatmap_scattering_img = np.asarray(heatmap_img[int(p), :, :])
            elif j == 1:
                heatmap_scattering_img = np.asarray(heatmap_img[:, int(p), :])
            elif j == 2:
                heatmap_scattering_img = np.asarray(heatmap_img[:, :, int(p)])
            heatmap_scattering_img = np.rot90(heatmap_scattering_img)
            heatmap_scattering_img[
                (heatmap_scattering_img < thresh_max) * (heatmap_scattering_img > thresh_min)] = np.nan

            if i == 0:
                # ax1.set_title(axis_type[j])
                ax1.set_ylabel(axis_type[j])
                # plt.ylabel(axis_type[j])
            # ax1.imshow(orig_scattering_img, cmap=cmap_orig, vmin=orig_vmin, vmax=orig_vmax)
            # im = ax1.imshow(heatmap_scattering_img, cmap=cmap_heatmap, alpha=alpha, vmin=positive_vmin, vmax=positive_vmax)
            im = ax1.imshow(heatmap_scattering_img, cmap=cmap_heatmap, alpha=alpha, vmin=vmin, vmax=vmax)
            ax1.set_yticks([])
            ax1.set_xticks([])
            ax1.spines['right'].set_visible(False)
            ax1.spines['top'].set_visible(False)
            ax1.spines['bottom'].set_visible(False)
            ax1.spines['left'].set_visible(False)
            axes.append(ax1)
            # ax1.axis('off')
            del heatmap_scattering_img

    # (left, bottom, width, height)
    cax = plt.axes([0.95, 0.1, 0.01, 0.8])
    cbar = fig.colorbar(im, ax=axes, extend='both', cax=cax)

    cbar.set_ticks(np.array((vmin, thresh_min, thresh_max, vmax)))
    cbar.set_ticklabels(["%.2f" % (vmin), "%.2f" % (thresh_min), "%.2f" % (thresh_max), "%.2f" % (vmax)])
    # plt.subplots_adjust(bottom=0.1, right=0.6, top=0.9, left=0.5)

    plt.tight_layout()
    plt.savefig(save_dir, dpi = 600)
    plt.close('all')




def data_augmentation(datas, cur_epoch):
    list_cropping_info = None
    translation_list = None
    if fst.flag_translation == True:
        flag_rand_trans = np.random.randint(1000)
        if fst.flag_translation_ratio != 0:
            if flag_rand_trans % fst.flag_translation_ratio == 0:
                pass
            else:
                size_of_translation = st.size_translation
                translation_list = np.random.randint(size_of_translation, size=(datas.size()[0], 3)) - size_of_translation // 2
                for batch_i in range(translation_list.shape[0]):
                    for axis_i in range(translation_list.shape[1]):
                        datas[batch_i][0] = ut.push_tensor(datas[batch_i][0], translation_list[batch_i][axis_i], axis_i)
        else:
            if fst.flag_translation_ratio_2 == True:
                size_of_translation = st.size_translation - int(cur_epoch // (st.epoch / st.size_translation))
                
            else:
                size_of_translation = st.size_translation
            
            if size_of_translation > 1:
                if fst.flag_translation_subject_wise == True:
                    translation_list = np.random.randint(size_of_translation, size=(datas.size()[0], 3)) - size_of_translation // 2
                    for batch_i in range(translation_list.shape[0]):
                        for axis_i in range(translation_list.shape[1]):
                            datas[batch_i][0] = ut.push_tensor(datas[batch_i][0], translation_list[batch_i][axis_i], axis_i)
                else:
                    translation_list = np.random.randint(size_of_translation, size=(3)) - size_of_translation // 2
                    translation_list = np.tile(translation_list, (datas.size()[0], 1))
                    for batch_i in range(translation_list.shape[0]):
                        for axis_i in range(translation_list.shape[1]):
                            datas[batch_i][0] = ut.push_tensor(datas[batch_i][0], translation_list[batch_i][axis_i], axis_i)
    if fst.flag_cropping == True:
        flag_rand_crop = np.random.randint(1000)
        if fst.flag_crop_ratio != 0:
            if flag_rand_crop % fst.flag_crop_ratio == 0:
                pass
            else:
                datas, list_cropping_info = ut.crop_tensor(datas)
        else:
            datas, list_cropping_info= ut.crop_tensor(datas)

    dict_result = {
        "datas": datas,
        "translation_list": translation_list,
        "list_cropping_info": list_cropping_info,
    }
    return dict_result

def plot_training_info_1(fold, dir_pyplot, EMS, flag = 'minmax', flag_match = False):
    """ plot the chat"""
    """ train loss """
    y_list = []
    y_list_name = []
    y_list.append(EMS.train_loss)
    y_list_name.append('train loss')
    ut.plot_list_v1(EMS.train_step, y_list, title='train loss', n_xlabel='step', n_ylabel=y_list_name,
                    save_dir=dir_pyplot, file_name='/fold_{0}_train_loss'.format(fold), flag=flag, flag_match=flag_match)
    del y_list, y_list_name

    """ train aux loss """
    tmp_flag = False
    y_list = []
    y_list_name = []
    for tmp_i in range(len(EMS.train_aux_loss)):
        if len(EMS.train_aux_loss[tmp_i]) == len(EMS.train_step):
            y_list.append(EMS.train_aux_loss[tmp_i])
            y_list_name.append('aux loss {}'.format(tmp_i + 1))
            tmp_flag = True
    if tmp_flag == True:
        ut.plot_list_v1(EMS.train_step, y_list, title='train aux loss', n_xlabel='step', n_ylabel=y_list_name,
                        save_dir=dir_pyplot, file_name='/fold_{0}_train_aux_loss'.format(fold), flag=flag, flag_match=flag_match)
    del y_list, y_list_name


    """ val test acc """
    tmp_flag = False
    y_list = []
    y_list_name = []
    if len(EMS.dict_val_metric['val_acc']) == len(EMS.val_step):
        y_list.append(EMS.dict_val_metric['val_acc'])
        y_list_name.append('val acc')
        tmp_flag = True
    if len(EMS.test_acc) == len(EMS.val_step):
        y_list.append(EMS.test_acc)
        y_list_name.append('test acc')
        tmp_flag = True
    if tmp_flag == True:
        ut.plot_list_v2(EMS.val_step, y_list, title='Val Test Acc Img', n_xlabel='step', n_ylabel=y_list_name,
                        save_dir=dir_pyplot, file_name='/fold_{0}_val_test_acc'.format(fold), flag=flag)
    del y_list, y_list_name

    #TODO : ROI acc
    """ val test aux acc"""
    for tmp_i in range(len(st.lambda_aux)):
        tmp_flag = False
        y_list = []
        y_list_name = []
        if len(EMS.dict_val_metric['val_acc_aux'][tmp_i]) == len(EMS.val_step):
            y_list.append(EMS.dict_val_metric['val_acc_aux'][tmp_i])
            y_list_name.append('val acc {}'.format(st.lambda_aux_name[tmp_i]))
            tmp_flag = True
        if len(EMS.test_acc_aux[tmp_i]) == len(EMS.val_step):
            y_list.append(EMS.test_acc_aux[tmp_i])
            y_list_name.append('test acc {}'.format(st.lambda_aux_name[tmp_i]))
            tmp_flag = True
        if tmp_flag == True:
            ut.plot_list_v2(EMS.val_step, y_list, title='Val Test Acc {}'.format(st.lambda_aux_name[tmp_i]), n_xlabel='step', n_ylabel=y_list_name,
                            save_dir=dir_pyplot, file_name='/fold_{0}_val_test_acc_{1}'.format(fold, st.lambda_aux_name[tmp_i]), flag=flag)

    """ val test loss """
    tmp_flag = False
    y_list = []
    y_list_name = []
    if len(EMS.dict_val_metric['val_loss']) == len(EMS.val_step):
        y_list.append(EMS.dict_val_metric['val_loss'])
        y_list_name.append('val loss')
        tmp_flag = True
    if len(EMS.test_loss) == len(EMS.val_step):
        y_list.append(EMS.test_loss)
        y_list_name.append('test loss')
        tmp_flag = True
    if tmp_flag == True:
        ut.plot_list_v2(EMS.val_step, y_list, title='Val Test Loss', n_xlabel='step', n_ylabel=y_list_name,
                    save_dir=dir_pyplot, file_name='/fold_{0}_val_test_loss'.format(fold), flag=flag)
    del y_list, y_list_name

    """ val test loss age"""
    tmp_flag = False
    y_list = []
    y_list_name = []
    if len(EMS.dict_val_metric['val_loss_age']) == len(EMS.val_step):
        y_list.append(EMS.dict_val_metric['val_loss_age'])
        y_list_name.append('val loss age')
        tmp_flag = True
    if len(EMS.test_loss_age) == len(EMS.val_step):
        y_list.append(EMS.test_loss_age)
        y_list_name.append('test loss age')
        tmp_flag = True
    if tmp_flag == True:
        ut.plot_list_v2(EMS.val_step, y_list, title='Val Test loss age', n_xlabel='step', n_ylabel=y_list_name,
                    save_dir=dir_pyplot, file_name='/fold_{0}_val_test_loss_age'.format(fold), flag=flag)
    del y_list, y_list_name

    """ learning rate """
    y_list = []
    y_list_name = []
    y_list.append(EMS.LR)
    y_list_name.append('learning rate')
    ut.plot_list_v1(EMS.val_step, y_list, title='Learning rate', n_xlabel='step', n_ylabel=y_list_name,
                    save_dir=dir_pyplot, file_name='/fold_{0}_Learning_rate'.format(fold), flag=flag, flag_match=flag_match)
    del y_list, y_list_name

    """ train acc """
    if fst.flag_print_trainAcc == True:
        y_list = []
        y_list_name = ['train acc']
        y_list.append(EMS.train_acc)
        ut.plot_list_v1(EMS.train_step, y_list, title='train acc', n_xlabel='step', n_ylabel=y_list_name,
                        save_dir=dir_pyplot, file_name='/fold_{0}_train_acc'.format(fold), flag=flag, flag_match=flag_match)
        del y_list, y_list_name



def eval_regression_model(config, loader, model, criterion):
    """ loader"""
    MAE_loss = 0
    criterion_MSE = nn.MSELoss()
    model.eval()
    torch.cuda.empty_cache()
    predict_result = []
    groundTruth=[]
    with torch.no_grad():
        for datas, labels, alabels, mlabel in loader:
            """ input"""
            datas = Variable(datas).cuda()
            labels = Variable(labels.long()).cuda()
            alabels = Variable(alabels.float()).cuda()

            """ get the output, logit"""
            model.eval()

            """ forward propagation """
            dict_result = model(datas)
            output_regression= dict_result['preds']
            predict_result.append(output_regression.data.cpu().numpy().squeeze())
            groundTruth.append(alabels.data.cpu().numpy().squeeze())

            """ calculate the loss """
            loss_1 = criterion(output_regression.squeeze(), alabels.squeeze())
            loss = loss_1
            MAE_loss += loss.data.cpu().numpy()

    total = len(loader.dataset)
    MAE_loss = MAE_loss / total
    predict_result_np = np.hstack(predict_result)
    groundTruth_np = np.hstack(groundTruth)
    r2 = r2_score(y_true=groundTruth_np, y_pred=predict_result_np )
    rmse = math.sqrt(mean_squared_error(y_true=groundTruth_np, y_pred=predict_result_np))
    dict_result = {
        "Loss": MAE_loss,
        "MAE": MAE_loss,
        "RMSE": rmse,
        "R_squared": r2,
    }
    return dict_result


def eval_regression_model_pretrained(config, fold, loader, model_1, model_2, criterion):
    """ loader"""
    np_loss = 0
    model_1.eval()
    model_2.eval()
    torch.cuda.empty_cache()
    with torch.no_grad():
        for datas, labels, alabels, mlabel in loader:

            """ input"""
            datas = Variable(datas).cuda()

            """ forward propagation """
            dict_result = model_1(datas)
            logitMap = dict_result['logitMap']

            data_2 = logitMap.detach()
            dict_result = model_2(data_2)
            f_logitMap = dict_result['logitMap']

            """ calculate the loss """
            loss = criterion(logitMap, f_logitMap)
            np_loss += loss.data.cpu().numpy()*f_logitMap.size(0)

    total = len(loader.dataset)
    np_loss = np_loss / total

    dict_result = {
        "Loss": np_loss,
    }
    return dict_result

def eval_multi_task_model(config, fold, loader, model, criterion_cls, criterion_L1, confusion_save_info = None):
    """ loader"""
    correct = 0
    list_aux_acc = [0] * len(st.lambda_aux)
    loss_np = 0
    age_total = 0
    loss_np_age = 0
    model.eval()
    torch.cuda.empty_cache()
    predict_result = []
    groundTruth_result = []
    predict_result_age = []
    groundTruth_age = []
    count = 0
    # MAE_loss = 0
    with torch.no_grad():
        for data_batch in loader:
            count +=1

            """ input"""
            datas = Variable(data_batch['data'].float()).cuda()
            labels = Variable(data_batch['label'].float()).cuda()
            labels_age = Variable(data_batch['age'].float()).cuda()
            labels_MMSE = Variable(data_batch['MMSE'].float()).cuda()

            """ encode age label """
            # labels_age = ((labels_age - 50) * 10).int()
            # age_encode = torch.zeros(labels_age.size(0), 9).cuda()
            # for tmp_i in range(9):
            #     age_encode[:, tmp_i] = labels_age % 2
            #     labels_age = labels_age // 2
            # age_encode_label = age_encode.clone()

            """ minmax norm """
            if st.list_data_norm_type[st.data_norm_type_num] == 'minmax':
                tmp_datas = datas.view(datas.size(0), -1)
                tmp_datas -= tmp_datas.min(1, keepdim=True)[0]
                tmp_datas /= tmp_datas.max(1, keepdim=True)[0]
                datas = tmp_datas.view_as(datas)



            """ get the output, logit"""
            model.eval()

            """ forward propagation """
            dict_result = model(datas, labels_age.clone())
            output_1 = dict_result['logits']
            output_2 = dict_result['Aux_logits']
            output_3 = dict_result['logitMap']

            """ loss 1 """
            loss_list_1 = []
            s_labels = ut.smooth_one_hot(labels, config.num_classes, smoothing=st.smoothing_img)
            loss_2 = criterion_cls(output_1, s_labels) * st.lambda_major[0] * output_1.size(0)
            loss_list_1.append(loss_2)

            """ loss 2 """
            loss_list_2 = []
            s_labels = labels_age
            loss_2 = criterion_L1(output_2[0], s_labels)
            loss_list_2.append(loss_2)

            """ add the loss """
            loss = sum(loss_list_1)
            loss_np += loss.data.cpu().numpy()

            if len(loss_list_2) != 0:
                loss_age = sum(loss_list_2)
                loss_np_age += loss_age.data.cpu().numpy()


            pred = torch.round(output_1)
            correct += pred.eq(labels.view_as(pred)).sum().cpu().item()

            """ stack ground truth and prediction """
            predict_result.append(pred.data.cpu().numpy().squeeze())
            groundTruth_result.append(np.hstack(labels.data.cpu().numpy()).squeeze())

            predict_result_age.append(output_2[0].data.cpu().numpy().squeeze())
            groundTruth_age.append(labels_age.data.cpu().numpy().squeeze())

    """ stacking """
    disease_groundTruth = np.hstack(groundTruth_result)
    disease_predict_result = np.hstack(predict_result)
    age_groundTruth = np.hstack(groundTruth_age)
    age_predict_result = np.hstack(predict_result_age)

    """ plot the confusion matrix """
    if confusion_save_info != None:
        """ plot and save the confusion matrix """
        class_names = np.zeros(len(st.list_selected_for_train), dtype="U10")
        for i in range (len(st.list_selected_for_train)):
            class_names[i] = st.list_selected_for_train[i]

        ut.plot_confusion_matrix(disease_groundTruth, disease_predict_result, classes=class_names, f_dir=confusion_save_info[1],
                                 f_name='/confusion_fold_' + str(confusion_save_info[0]) + '.png',
                                 title='Confusion matrix, without normalization')

    if len(st.list_selected_for_train) == 2:
        """ Confusion matrix , Accuracy, sensitvity and specificity """
        cm1 = confusion_matrix(disease_groundTruth, disease_predict_result)

        """ calculate the accuracy """
        total1 = sum(sum(cm1))
        accuracy1 = (cm1[0, 0] + cm1[1, 1]) / total1
        sensitivity1 = cm1[0, 0] / (cm1[0, 0] + cm1[0, 1])
        specificity1 = cm1[1, 1] / (cm1[1, 0] + cm1[1, 1])
        fpr, tpr, thresholds = metrics.roc_curve(disease_groundTruth, disease_predict_result, pos_label=1)
        AUC = metrics.auc(fpr, tpr)

    else:
        accuracy1 = correct / len(loader.dataset)
        sensitivity1 = 0
        specificity1 = 0
        AUC = 0

    # MAE_loss = MAE_loss / age_total
    loss_np_age = loss_np_age / len(loader.dataset)
    predict_result_np = np.hstack(age_predict_result)
    groundTruth_np = np.hstack(age_groundTruth)
    r2 = r2_score(y_true=groundTruth_np, y_pred=predict_result_np)
    rmse = math.sqrt(mean_squared_error(y_true=groundTruth_np, y_pred=predict_result_np))

    """ calculate the accuracy and print """
    total = len(loader.dataset)
    loss_np = loss_np / total
    dict_result = {
        "Loss": loss_np,
        "loss_age": loss_np_age,
        "Acc": accuracy1,
        "Sen": sensitivity1,
        "Spe": specificity1,
        "AUC": AUC,
        "MAE": loss_np_age,
        # "RMSE": None,
        "RMSE": rmse,
        # "R_squared": None,
        "R_squared": r2,
    }
    return dict_result


def estimate_biological_age(age, MMSE, lambda_disease_factor):
    """ params """
    age_normalization_factor = 1 / 70

    if st.list_age_estimating_function[st.selected_function] == 'linear_1':
        # linear function
        age = age + (30 - MMSE) * lambda_disease_factor

    elif st.list_age_estimating_function[st.selected_function] == 'linear_1_with_age':
        # linear function with age
        age = age + (30 - MMSE) * lambda_disease_factor * age * age_normalization_factor

    elif st.list_age_estimating_function[st.selected_function] == 'sigmoid_1':
        # using sigmoid
        w_x = -0.1
        b_x = 15
        w_y = 30
        b_y = -5
        age = age + ((1 / (1 + np.exp(-(w_x * (MMSE - b_x))))) * w_y + b_y) * lambda_disease_factor

    elif st.list_age_estimating_function[st.selected_function] == 'sigmoid_1_with_age':
        # using sigmoid
        w_x = -0.1
        b_x = 15
        w_y = 30
        b_y = -5
        age = age + ((1 / (1 + np.exp(-(w_x * (MMSE - b_x))))) * w_y + b_y) * age * age_normalization_factor * lambda_disease_factor
    elif st.list_age_estimating_function[st.selected_function] == 'sqrt_1':
        # using sigmoid
        w = 4.5
        b = 30
        age = age + (w * (np.sqrt(-MMSE + b))) * lambda_disease_factor
    elif st.list_age_estimating_function[st.selected_function] == 'quadratic_1':
        # using sigmoid
        w = 0.05
        b = 30
        age = age + (w * np.power(MMSE-b, 2)) * lambda_disease_factor

    elif st.list_age_estimating_function[st.selected_function] == 'none':
        pass

    elif st.list_age_estimating_function[st.selected_function] == 'constant':
        age = age + 3 * lambda_disease_factor


    else:
        assert False, "the age estimation function has not been applied!"

    return age

def plot_logit_on_2D(logit_0, logit_1, gt, pred, save_dir = './', save_file_name = 'test', fold=None):
    x = logit_0
    y = logit_1
    ## TODO: scatter logit
    fig, ax = plt.subplots()
    ax.scatter(x, y, s=1)
    t_min = int(min([x.reshape(-1).min(), y.reshape(-1).min()]) * 1.2)
    t_max = int(max([x.reshape(-1).max(), y.reshape(-1).max()]) * 1.2)
    ax.plot(range(t_min, t_max), range(t_min, t_max))
    ax.set_xlim([t_min, t_max])
    ax.set_ylim([t_min, t_max])
    ax.grid(True)
    plt.axis('square')

    f_save_dir = save_dir + '/fold_{}/gt_{}_pred_{}'.format(fold, gt, pred)
    make_dir(f_save_dir)
    plt.savefig(f_save_dir + '/{}'.format(save_file_name))
    plt.close('all')


def scatter_AD_logit(AD_logit_img, featureMaps, sigma_factor=2, save_dir = './', save_file_name= '', fold=1, pred = None, label = None):
    """
    :param AD_logit_img: # w, h, d
    :param featureMaps: # f, w, h, d
    :return:
    """
    tmp = AD_logit_img
    tmp_mean = torch.mean(tmp, dim=(0, 1, 2), keepdim=False)
    tmp_std = torch.std(tmp, dim=(0, 1, 2), keepdim=False)
    tmp_z = (tmp - tmp_mean) / tmp_std
    tmp_mask = (tmp_z > sigma_factor).float().squeeze().cpu().numpy()
    orig = torch.mean(featureMaps, dim=1).squeeze().cpu().numpy()
    f_save_dir = save_dir + '/fold_{}/gt_{}_pred_{}'.format(fold, label, pred)
    ut.make_dir(dir=f_save_dir, flag_rm=False)
    ut.plot_heatmap_with_overlay(orig_img=orig, heatmap_img=tmp_mask,
                                 save_dir=f_save_dir + '/{}'.format(save_file_name), fig_title='zero_mean_unit_var_{}'.format(sigma_factor), thresh=0.2,
                                 percentile=1)

def scatter_AD_logit_with_mask(AD_logit_img, featureMaps, sigma_factor=2, save_dir = './', save_file_name= '', fold=1, pred = None, label = None):
    """
    :param AD_logit_img: # w, h, d
    :param featureMaps: # f, w, h, d
    :return:
    """
    tmp = AD_logit_img  # batch, 18, 23, 17

    ## consider positive only
    tmp_w = tmp > 0 # mask

    tmp_mean = torch.sum(tmp * tmp_w.float()) / torch.sum(tmp_w)
    tmp_std = torch.sqrt(torch.sum((tmp - tmp_mean) ** 2 * tmp_w.float()) / torch.sum(tmp_w))

    # tmp_z = (tmp - tmp_mean) / tmp_std
    tmp_mask = (tmp > tmp_mean + sigma_factor * tmp_std).float().squeeze().cpu().numpy()
    orig = torch.mean(featureMaps, dim=1).squeeze().cpu().numpy()
    f_save_dir = save_dir + '/fold_{}/gt_{}_pred_{}'.format(fold, label, pred)
    ut.make_dir(dir=f_save_dir, flag_rm=False)
    ut.plot_heatmap_with_overlay(orig_img=orig, heatmap_img=tmp_mask,
                                 save_dir=f_save_dir + '/{}'.format(save_file_name), fig_title='masked_with_positive_AD_logits_sigma_{}'.format(sigma_factor), thresh=0.2,
                                 percentile=1)


def plot_hist_patch_level_AD_logit(output_logitMap, count, fold, labels, pred, save_dir, file_name):
    fig, ax = plt.subplots()
    tmp = output_logitMap[:, 1].squeeze().view(-1).data.cpu().numpy()
    y, x, patches = plt.hist(tmp, bins=10, facecolor='blue', alpha=0.5)
    for i, v in enumerate(y):
        ax.text(x[i], v + 5, str(v), color='black')
    plt.xlabel('AD logit')
    plt.ylabel('# of patches')
    plt.title('Histogram of Patch-level AD logit distribution in sample{} fold{} (gt:{} pred:{})'.format(count, fold, labels.squeeze().data.cpu().numpy(), pred.squeeze().data.cpu().numpy()), fontsize= 10)
    make_dir(save_dir)
    plt.savefig(save_dir + file_name)
    plt.close('all')

def plot_violin(data, save_dir, file_name):
    ## data = [NC, AD] [9, 17, 33] [sample]

    heights = []
    widths = []
    for i_tmp in range(len(data[0])):
        heights.append(3)
    for i_tmp in range(len(data) * len(data[0][0])):
        widths.append(2)

    fig = plt.figure(figsize=(len(widths) * 3, len(heights) * 4))
    plt.rcParams.update({'font.size': 20})

    gs = gridspec.GridSpec(nrows=len(heights),  # row
                           ncols=len(widths),  # col
                           height_ratios=heights,
                           width_ratios=widths
                           )
    # colors = ["crimson", "indigo", "limegreen", "gold"]


    for j_row in range(len(heights)):
        min=np.inf
        max = -np.inf
        for i_col in range(len(widths)):
            index_1 = i_col // len(data[0][0])
            index_2 = j_row
            index_3 = i_col % len(data[0][0])
            y = data[index_1][index_2][index_3]

            if y.max() > max:
                max =y.max()
            if y.min() < min:
                min= y.min()
        margin = 1.5
        min *= margin
        max *= margin
        if -min > max:
            abs_max = -min
        else:
            abs_max = max
        for i_col in range(len(widths)):
            ax1 = fig.add_subplot(gs[j_row, i_col])
            index_1 = i_col // len(data[0][0])
            index_2 = j_row
            index_3 = i_col % len(data[0][0])
            y = data[index_1][index_2][index_3]
            if index_1 == 0:
                color = 'skyblue'
            else:
                color = 'darkcyan'
            ax1 = sns.violinplot(y=y.reshape(-1), color=color,fliersize=15, linewidth=2)
            # ax1 = sns.violinplot(y=y.reshape(-1),bw='scott', color=color,fliersize=15, linewidth=2)
            # ax1 = sns.violinplot(y=y.reshape(-1), bw='silverman', color=color, fliersize=15, linewidth=2)
            if i_col == 0 :
                # ax1.set_ylabel('Patch Size : 9')
                pass
            if j_row == 0 :
                for tmp_i in range(len(widths)):
                    if i_col == tmp_i:
                        ax1.set_xlabel('Sample {}'.format(tmp_i+1))
                        ax1.xaxis.set_label_position('top')


            ax1.text(0, abs_max* 0.8, '\u03C3 : '+'%.2f'%(np.std(y.reshape(-1))), horizontalalignment='center', fontsize=20)
            ax1.set_ylim([-abs_max,abs_max])
            ax1.grid(True)
    gs.tight_layout(fig)


    make_dir(save_dir)
    plt.savefig(save_dir + file_name)
    plt.close('all')


def plot_group_violin(data, save_dir, file_name):
    ## data = [NC, AD] [9, 17, 33] [sample]
    NC_9 = np.vstack(data[0][0]).mean(axis=0).reshape(-1)
    NC_17 = np.vstack(data[0][1]).mean(axis=0).reshape(-1)
    NC_33 = np.vstack(data[0][2]).mean(axis=0).reshape(-1)

    AD_9 = np.vstack(data[1][0]).mean(axis=0).reshape(-1)
    AD_17 = np.vstack(data[1][1]).mean(axis=0).reshape(-1)
    AD_33 = np.vstack(data[1][2]).mean(axis=0).reshape(-1)


    heights = [2]
    widths = []
    for i_tmp in range(len(data[0])):
        widths.append(3)

    fig = plt.figure(figsize=(len(widths) * 8, len(heights) * 6))
    params = {
        'font.size': 30,
        'legend.fontsize': 20,
    }
    plt.rcParams.update(params)

    gs = gridspec.GridSpec(nrows=len(heights),  # row
                           ncols=len(widths),  # col
                           height_ratios=heights,
                           width_ratios=widths
                           )
    # colors = ["crimson", "indigo", "limegreen", "gold"]

    for j_row in range(len(heights)):
        min = np.inf
        max = -np.inf

        for i_col in range(len(widths)):
            ax1 = fig.add_subplot(gs[j_row, i_col])

            if i_col == 0 :
                t_data_0 = NC_9
                t_data_1 = AD_9
            elif i_col == 1 :
                t_data_0 = NC_17
                t_data_1 = AD_17
            elif i_col == 2:
                t_data_0 = NC_33
                t_data_1 = AD_33


            tmp_a = ['NC' for _ in range(t_data_0.shape[0])]
            tmp_b = ['AD' for _ in range(t_data_1.shape[0])]
            class_data = tmp_a + tmp_b
            logit_data = np.hstack([t_data_0, t_data_1])
            t_data = {
                'AD logit': logit_data,
                'Class': class_data,
                'Num of patches': '',
            }
            frame = DataFrame(t_data)

            min = logit_data.min()
            max = logit_data.max()
            margin = 1.2
            min *= margin
            max *= margin
            if -min > max:
                abs_max = -min
            else:
                abs_max = max

            color = ['skyblue', 'darkcyan']
            ax1 = sns.violinplot(x='AD logit' ,y='Num of patches', hue='Class', palette={"NC": color[0], "AD": color[1]}, data=frame, fliersize=15, linewidth=2, split=True)
            # ax1.set_xlim([-abs_max, abs_max])
            ax1.set_xlim([-80, 80])
            ax1.grid(True)



            # ax1.text(0, abs_max * 0.8, '\u03C3 : ' + '%.2f' % (np.std(y.reshape(-1))), horizontalalignment='center',
            #          fontsize=20)
            # ax1.set_ylim([-abs_max, abs_max])
            ax1.grid(True)
    gs.tight_layout(fig)

    make_dir(save_dir)
    plt.savefig(save_dir + file_name)
    plt.close('all')


def plot_patch_boundary(list_orig_img, list_heatmap_img, save_dir, fig_title = 'patch_boundary', thresh=0.5, patch_size = [9, 17, 33]):
    ## list_heatmap_img [3][3][1]

    ## TODO : input
    # loop for the subjects
    patch_size = patch_size
    stride_between_patches = st.patch_stride
    reduce_size_axis_1 = 5
    reduce_size_axis_2 = 5
    reduce_size_axis_3_a = 0
    reduce_size_axis_3_b = 20
    for i_sub in range(len(list_orig_img)):
        print('i_sub : {}'.format(i_sub))
        # loop for patch size
        percentile = 0.2
        ## TODO :  Figure
        orig_img = list_orig_img[i_sub][0, 0]  # 64, 64, 64
        orig_vmax = np.percentile(orig_img, 100 - percentile)
        orig_vmin = np.percentile(orig_img, percentile)
        print(orig_vmin, orig_vmax)
        shape = orig_img.shape
        list_interval = []
        for j in range(3):
            tmp_list = []
            if j == 0 or j == 1:
                a = 30
                b = 71
            else:
                a = 20
                b = 61
            for i in np.arange(a, b, 10):
                tmp_list.append(int(np.percentile(np.arange(0, shape[j]), i)))
            list_interval.append(np.hstack(tmp_list))

        axis_type = ['Sagittal', 'Coronal', 'Axial']

        ## TODO : heatmap generate
        list_mask = []
        list_index = []
        list_heatmap = []
        for i_patch in range(len(list_heatmap_img)):

            fig = plt.figure(figsize=(list_interval[0].shape[0] * 2, len(axis_type) * 2))
            plt.rcParams.update({'font.size': 20})
            # fig.suptitle(fig_title, fontsize=20)

            heights = [1] * len(axis_type)
            widths = [10] * (list_interval[0].shape[0])
            gs = gridspec.GridSpec(nrows=len(heights),  # row
                                   ncols=len(widths),
                                   height_ratios=heights,
                                   width_ratios=widths,
                                   hspace=0.0,
                                   wspace=0.0,
                                   )
            cmap_orig = plt.get_cmap('Greys')
            for j, q in enumerate(axis_type):
                for i, p in enumerate(list_interval[j]):
                    ax1 = fig.add_subplot(gs[j, i])
                    if j == 0:
                        orig_scattering_img = np.asarray(orig_img[int(p), reduce_size_axis_2:-reduce_size_axis_2, reduce_size_axis_3_a:-reduce_size_axis_3_b])
                    elif j == 1:
                        orig_scattering_img = np.asarray(orig_img[reduce_size_axis_1:-reduce_size_axis_1, int(p), reduce_size_axis_3_a:-reduce_size_axis_3_b])
                    elif j == 2:
                        orig_scattering_img = np.asarray(orig_img[reduce_size_axis_1:-reduce_size_axis_1, reduce_size_axis_2:-reduce_size_axis_2, int(p)])
                    orig_scattering_img = np.rot90(orig_scattering_img)
                    ax1.imshow(orig_scattering_img, cmap=cmap_orig, vmin=orig_vmin, vmax=orig_vmax)
                    ax1.set_yticks([])
                    ax1.set_xticks([])
                    ax1.spines['right'].set_visible(False)
                    ax1.spines['top'].set_visible(False)
                    ax1.spines['bottom'].set_visible(False)
                    ax1.spines['left'].set_visible(False)
                    ax1.axis('off')

            print('i_patch : {}'.format(i_patch))
            list_mask.append([])
            list_index.append([])
            list_heatmap.append([])

            ## TODO :  topk
            t_heatmap_img = list_heatmap_img[i_patch][i_sub]  # 1, 7, 7, 7
            index = np.argsort(t_heatmap_img, axis=None)

            n_k = 5
            topk = t_heatmap_img.reshape(-1)[index[-n_k:]]
            for i_top in range(n_k):
                print('i_top : {}'.format(i_top))
                list_heatmap[i_patch].append(np.zeros_like(list_orig_img[0].squeeze()))
                list_mask[i_patch].append(np.where(t_heatmap_img == topk[i_top], 1, 0))
                list_index[i_patch].append(np.argwhere(list_mask[i_patch][i_top] == 1).squeeze()[1:])
                list_index[i_patch][i_top] = list_index[i_patch][i_top] * stride_between_patches + patch_size[
                    i_patch] // 2
                ##TODO : generate mask...
                list_heatmap[i_patch][i_top][
                        list_index[i_patch][i_top][0] - patch_size[i_patch] // 2:
                        list_index[i_patch][i_top][0] + patch_size[i_patch] // 2 + 1,

                        list_index[i_patch][i_top][1] - patch_size[i_patch] // 2:
                        list_index[i_patch][i_top][1] + patch_size[i_patch] // 2 + 1,

                        list_index[i_patch][i_top][2] - patch_size[i_patch] // 2:
                        list_index[i_patch][i_top][2] + patch_size[i_patch] // 2 + 1,
                ] = 1

                # list_heatmap[i_patch][i_top][
                #         list_index[i_patch][i_top][0] - patch_size[i_patch] // 2 +1:
                #         list_index[i_patch][i_top][0] + patch_size[i_patch] // 2,
                #
                #         list_index[i_patch][i_top][1] - patch_size[i_patch] // 2+1:
                #         list_index[i_patch][i_top][1] + patch_size[i_patch] // 2,
                #
                #         list_index[i_patch][i_top][2] - patch_size[i_patch] // 2+1:
                #         list_index[i_patch][i_top][2] + patch_size[i_patch] // 2,
                # ] = 0

                ## TODO : plot heatmap
                # cmap_heatmap = plt.get_cmap('Reds')
                # cmap_heatmap = plt.get_cmap('coolwarm')

                # color_list = [['lightcoral', 'firebrick'],
                #               ['mediumseagreen', 'darkgreen'],
                #               ['skyblue', 'navy'],
                #               ]
                # cmap_heatmap = mcolors.ListedColormap(['white', color_list[i_patch][i_top]])
                color_list = ['Reds', 'Greens', 'Blues']

                tmp_n_k = n_k *2
                tmp_array = (np.arange(tmp_n_k) / tmp_n_k) + (1 - max((np.arange(tmp_n_k) / tmp_n_k)))
                color = plt.get_cmap(color_list[i_patch])(tmp_array[i_top + n_k])
                cmap_heatmap = mcolors.ListedColormap(['white', color])
                percentile = 0
                heatmap_img = list_heatmap[i_patch][i_top]  # 64, 64, 64
                vmax = np.percentile(heatmap_img, 100-percentile)
                vmin = np.percentile(heatmap_img, percentile)
                if np.abs(vmax) > np.abs(vmin):
                    vmax = np.abs(vmax)
                    vmin = -np.abs(vmax)
                else:
                    vmax = np.abs(vmin)
                    vmin = -np.abs(vmin)

                thresh_max = vmax * thresh
                thresh_min = vmin * thresh


                alpha = 0.3
                axes = []
                for j, q in enumerate(axis_type):
                    for i, p in enumerate(list_interval[j]):

                        # ax1 = fig.add_subplot(gs[j, i])
                        ax1 = plt.subplot(gs[j, i])
                        # ax1 = (gs[j, i])
                        if j == 0:
                            heatmap_scattering_img = np.asarray(heatmap_img[int(p), reduce_size_axis_2:-reduce_size_axis_2, reduce_size_axis_3_a:-reduce_size_axis_3_b])
                        elif j == 1:
                            heatmap_scattering_img = np.asarray(heatmap_img[reduce_size_axis_1:-reduce_size_axis_1, int(p), reduce_size_axis_3_a:-reduce_size_axis_3_b])
                        elif j == 2:
                            heatmap_scattering_img = np.asarray(heatmap_img[reduce_size_axis_1:-reduce_size_axis_1, reduce_size_axis_2:-reduce_size_axis_2, int(p)])


                        heatmap_scattering_img = np.rot90(heatmap_scattering_img)
                        heatmap_scattering_img[
                            (heatmap_scattering_img < thresh_max) * (heatmap_scattering_img > thresh_min)] = np.nan
                        # im = ax1.imshow(heatmap_scattering_img, cmap=cmap_heatmap, alpha=alpha, vmin=vmin, vmax=vmax)
                        im = ax1.imshow(heatmap_scattering_img, cmap=cmap_heatmap, alpha=alpha, vmin=vmin, vmax=vmax)
                        ax1.set_yticks([])
                        ax1.set_xticks([])
                        ax1.spines['right'].set_visible(False)
                        ax1.spines['top'].set_visible(False)
                        ax1.spines['bottom'].set_visible(False)
                        ax1.spines['left'].set_visible(False)
                        axes.append(ax1)

                    # ax1.axis('off')


            # (left, bottom, width, height)
            # cax = plt.axes([0.95, 0.1, 0.01, 0.8])
            # cbar = fig.colorbar(im, ax=axes, extend='both', cax=cax)

            # cbar.set_ticks(np.array((vmin, thresh_min, thresh_max, vmax)))
            # cbar.set_ticklabels(["%.2f" % (vmin), "%.2f" % (thresh_min), "%.2f" % (thresh_max), "%.2f" % (vmax)])
            # plt.subplots_adjust(bottom=0.1, right=0.6, top=0.9, left=0.5)

            plt.tight_layout()
            make_dir(save_dir)
            plt.savefig(save_dir + '/sub_{}_patch_{}'.format(i_sub, i_patch))
            plt.close('all')


import torch
from torch.optim.optimizer import Optimizer


class BayesianSGD(Optimizer):

    def __init__(self, params, momentum=0, dampening=0,
                 weight_decay=0, nesterov=False):
        if momentum < 0.0:
            raise ValueError("Invalid momentum value: {}".format(momentum))
        if weight_decay < 0.0:
            raise ValueError("Invalid weight_decay value: {}".format(weight_decay))

        defaults = dict(momentum=momentum, dampening=dampening,
                        weight_decay=weight_decay, nesterov=nesterov)
        if nesterov and (momentum <= 0 or dampening != 0):
            raise ValueError("Nesterov momentum requires a momentum and zero dampening")
        super(BayesianSGD, self).__init__(params, defaults)


    def __setstate__(self, state):
        super(BayesianSGD, self).__setstate__(state)

        for group in self.param_groups:
            group.setdefault('nesterov', False)

    def step(self, closure=None):
        """Performs a single optimization step.
        Arguments:
            closure (callable, optional): A closure that reevaluates the model
                and returns the loss.
        """

        loss = None
        if closure is not None:
            loss = closure()

        for group in self.param_groups:
            weight_decay = group['weight_decay']
            momentum = group['momentum']
            dampening = group['dampening']
            nesterov = group['nesterov']
            # uncertainty = group['uncertaintsy']

            for p in group['params']:
                if p.grad is None:
                    continue
                d_p = p.grad.data

                if weight_decay != 0:
                    d_p.add_(weight_decay, p.data)
                if momentum != 0:
                    param_state = self.state[p]
                    if 'momentum_buffer' not in param_state:
                        buf = param_state['momentum_buffer'] = torch.clone(d_p).detach()
                    else:
                        buf = param_state['momentum_buffer']
                        buf.mul_(momentum).add_(1 - dampening, d_p)
                    if nesterov:
                        d_p = d_p.add(momentum, buf)
                    else:
                        d_p = buf

                if isinstance(group['lr'], torch.Tensor):
                    p.data = p.data + torch.mul(-group['lr'].data, d_p)
                else:
                    p.data.add_(-group['lr'], d_p)

        return loss

def smooth_one_hot(targets, n_classes, smoothing = 0.0):
    assert 0 <= smoothing < 1
    with torch.no_grad():
        targets = torch.empty(size = (targets.size(0), n_classes)).cuda().fill_(smoothing / (n_classes-1)).scatter_(1, targets.long().data.unsqueeze(-1), 1-smoothing)
    return targets[:, -1]
